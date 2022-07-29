'''

Author       :  AnurajPilanku

Usecase      : EDI Aging

Code utility : Sent mail with table in mailbody

'''
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.application import MIMEApplication
from os.path import basename
import openpyxl
import pandas as pd
import sys
from datetime import date
import datetime

#cmd arguments

input=sys.argv[1]
greeting=sys.argv[2]#'''Hi All,'''
sentence=sys.argv[3]#'''PFB ageing ticket trend for EDI applications.(Note : Only reassigned tickets from EDI support group are considered)             #Refer the attached dump for further details'''

w=openpyxl.load_workbook(input+"\\"+"NGGcount.xlsx")#(r"\\acdev01\3M_CAC\EDI_Ageing\ediqueryoutput\engg.xlsx")
s=w.active
w1=openpyxl.load_workbook(input+"\\"+"Webmethodscount.xlsx")#(r"\\acdev01\3M_CAC\EDI_Ageing\ediqueryoutput\eweb.xlsx")
s1=w1.active
w2=openpyxl.load_workbook(input+"\\"+"Analystscount.xlsx")#(r"\\acdev01\3M_CAC\EDI_Ageing\ediqueryoutput\eanalysts.xlsx")
s2=w2.active
w3=openpyxl.load_workbook(input+"\\"+"Supportcount.xlsx")#(r"\\acdev01\3M_CAC\EDI_Ageing\ediqueryoutput\esupport.xlsx")
s3=w3.active
di=dict()
for i in range(2,s.max_row+1):
    di[s['A'+str(i)].value.strip()]=str(s['B'+str(i)].value)
    di[s1['A' + str(i)].value.strip()] = str(s1['B' + str(i)].value)
    di[s2['A' + str(i)].value.strip()] = str(s2['B' + str(i)].value)
    di[s3['A' + str(i)].value.strip()] = str(s3['B' + str(i)].value)
html_file = '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Title</title>
</head>
<body style="font-family:Cambria">
 <br/><img src='cid:image1'<br/>
  <br>
  <br>
  <br /><font face='Cambria'>'''+greeting+''' </a></font><br/>
  <br /><font face='Cambria'>'''+sentence+''' </a></font><br/>
  <br>
  <br>
<div style="overflow-x:auto;">
    <style>
        body{
        text-align:center;
        }
        table{
        border-collapse:collapse;}
        th,td{
        border: 1px solid black}
        th,td{
        padding:1px}
        .age{
        background-color:#CF9FFF;
        }
        .agenum{
        background-color:#FEF8DD;
        }
        .grandtotal{
        height:1px;
        padding: 0px;
        }
        .values{
        background-color:#D3D3D3;
        }
        .supportgroup{
        background-color:"#90EE90"
        }
    </style>
    </style>
        <table>
              <tr>
            <td class="grandtotal" colspan="7" bgcolor="#474227"  style="text-align:center;"> <p><font color="#FFFFFF">Grand Total</font></p></td>
        </tr>
        
        <tr bgcolor="#90EE90">
            <td colspan="2"  bgcolor="#C4A484" style="text-align:center;" >Status</td>
            <td class="values">Open</td>
            <td class="values">Pending Client</td>
            <td class="values">Pending Emergency Client</td>
            <td class="values">Pending Vendor</td>
            <td class="values">Work in Progress</td>
        </tr>
     
        <tr>
            <td class="age" rowspan="5">Aging</td>
            <td class="agenum"><=10</td>
            <td>'''+str(int(di["NggMSOpenCriticalLessThan10"])+int(di["NggMSOpenHighLessThan10"])+int(di["NggMSOpenAverageLessThan10"])+int(di["NggMSOpenLowLessThan10"])+int(di["NggNonMSOpenCriticalLessThan10"])+int(di["NggNonMSOpenHighLessThan10"])+int(di["NggNonMSOpenAverageLessThan10"])+int(di["NggNonMSOpenLowLessThan10"])+int(di["webNonMSOpenCriticalLessThan10"])+int(di["webNonMSOpenHighLessThan10"])+int(di["webNonMSOpenAverageLessThan10"])+int(di["webNonMSOpenLowLessThan10"])+int(di["webMSOpenCriticalLessThan10"])+int(di["webMSOpenHighLessThan10"])+int(di["webMSOpenAverageLessThan10"])+int(di["webMSOpenLowLessThan10"])+int(di["eanNonMSOpenCriticalLessThan10"])+int(di["eanNonMSOpenHighLessThan10"])+int(di["eanNonMSOpenAverageLessThan10"])+int(di["eanNonMSOpenLowLessThan10"])+int(di["eanMSOpenCriticalLessThan10"])+int(di["eanMSOpenHighLessThan10"])+int(di["eanMSOpenAverageLessThan10"])+int(di["eanMSOpenLowLessThan10"])+int(di["edsuMSOpenCriticalLessThan10"])+int(di["edsuMSOpenHighLessThan10"])+int(di["edsuMSOpenAverageLessThan10"])+int(di["edsuMSOpenLowLessThan10"])+int(di["edsuNonMSOpenCriticalLessThan10"])+int(di["edsuNonMSOpenHighLessThan10"])+int(di["edsuNonMSOpenAverageLessThan10"])+int(di["edsuNonMSOpenLowLessThan10"]))+'''</td>
            <td>'''+str(int(di["NggMSPending_ClientCriticalLessThan10"])+int(di["NggMSPending_ClientHighLessThan10"])+int(di["NggMSPending_ClientAverageLessThan10"])+int(di["NggMSPending_ClientLowLessThan10"])+int(di["NggNonMSPending_ClientCriticalLessThan10"])+int(di["NggNonMSPending_ClientHighLessThan10"])+int(di["NggNonMSPending_ClientAverageLessThan10"])+int(di["NggNonMSPending_ClientLowLessThan10"])+int(di["webNonMSPending_ClientCriticalLessThan10"])+int(di["webNonMSPending_ClientHighLessThan10"])+int(di["webNonMSPending_ClientAverageLessThan10"])+int(di["webNonMSPending_ClientLowLessThan10"])+int(di["webMSPending_ClientCriticalLessThan10"])+int(di["webMSPending_ClientHighLessThan10"])+int(di["webMSPending_ClientAverageLessThan10"])+int(di["webMSPending_ClientLowLessThan10"])+int(di["eanNonMSPending_ClientCriticalLessThan10"])+int(di["eanNonMSPending_ClientHighLessThan10"])+int(di["eanNonMSPending_ClientAverageLessThan10"])+int(di["eanNonMSPending_ClientLowLessThan10"])+int(di["eanMSPending_ClientCriticalLessThan10"])+int(di["eanMSPending_ClientHighLessThan10"])+int(di["eanMSPending_ClientAverageLessThan10"])+int(di["eanMSPending_ClientLowLessThan10"])+int(di["edsuMSPending_ClientCriticalLessThan10"])+int(di["edsuMSPending_ClientHighLessThan10"])+int(di["edsuMSPending_ClientAverageLessThan10"])+int(di["edsuMSPending_ClientLowLessThan10"])+int(di["edsuNonMSPending_ClientCriticalLessThan10"])+int(di["edsuNonMSPending_ClientHighLessThan10"])+int(di["edsuNonMSPending_ClientAverageLessThan10"])+int(di["edsuNonMSPending_ClientLowLessThan10"]))+'''</td>
            <td>'''+str(int(di["NggMSPending_Emergency_ClientCriticalLessThan10"])+int(di["NggMSPending_Emergency_ClientHighLessThan10"])+int(di["NggMSPending_Emergency_ClientAverageLessThan10"])+int(di["NggMSPending_Emergency_ClientLowLessThan10"])+int(di["NggNonMSPending_Emergency_ClientCriticalLessThan10"])+int(di["NggNonMSPending_Emergency_ClientHighLessThan10"])+int(di["NggNonMSPending_Emergency_ClientAverageLessThan10"])+int(di["NggNonMSPending_Emergency_ClientLowLessThan10"])+int(di["webNonMSPending_Emergency_ClientCriticalLessThan10"])+int(di["webNonMSPending_Emergency_ClientHighLessThan10"])+int(di["webNonMSPending_Emergency_ClientAverageLessThan10"])+int(di["webNonMSPending_Emergency_ClientLowLessThan10"])+int(di["webMSPending_Emergency_ClientCriticalLessThan10"])+int(di["webMSPending_Emergency_ClientHighLessThan10"])+int(di["webMSPending_Emergency_ClientAverageLessThan10"])+int(di["webMSPending_Emergency_ClientLowLessThan10"])+int(di["eanNonMSPending_Emergency_ClientCriticalLessThan10"])+int(di["eanNonMSPending_Emergency_ClientHighLessThan10"])+int(di["eanNonMSPending_Emergency_ClientAverageLessThan10"])+int(di["eanNonMSPending_Emergency_ClientLowLessThan10"])+int(di["eanMSPending_Emergency_ClientCriticalLessThan10"])+int(di["eanMSPending_Emergency_ClientHighLessThan10"])+int(di["eanMSPending_Emergency_ClientAverageLessThan10"])+int(di["eanMSPending_Emergency_ClientLowLessThan10"])+int(di["edsuMSPending_Emergency_ClientCriticalLessThan10"])+int(di["edsuMSPending_Emergency_ClientHighLessThan10"])+int(di["edsuMSPending_Emergency_ClientAverageLessThan10"])+int(di["edsuMSPending_Emergency_ClientLowLessThan10"])+int(di["edsuNonMSPending_Emergency_ClientCriticalLessThan10"])+int(di["edsuNonMSPending_Emergency_ClientHighLessThan10"])+int(di["edsuNonMSPending_Emergency_ClientAverageLessThan10"])+int(di["edsuNonMSPending_Emergency_ClientLowLessThan10"]))+'''</td>
            <td>'''+str(int(di["NggMSPending_VendorCriticalLessThan10"])+int(di["NggMSPending_VendorHighLessThan10"])+int(di["NggMSPending_VendorAverageLessThan10"])+int(di["NggMSPending_VendorLowLessThan10"])+int(di["NggNonMSPending_VendorCriticalLessThan10"])+int(di["NggNonMSPending_VendorHighLessThan10"])+int(di["NggNonMSPending_VendorAverageLessThan10"])+int(di["NggNonMSPending_VendorLowLessThan10"])+int(di["webNonMSPending_VendorCriticalLessThan10"])+int(di["webNonMSPending_VendorHighLessThan10"])+int(di["webNonMSPending_VendorAverageLessThan10"])+int(di["webNonMSPending_VendorLowLessThan10"])+int(di["webMSPending_VendorCriticalLessThan10"])+int(di["webMSPending_VendorHighLessThan10"])+int(di["webMSPending_VendorAverageLessThan10"])+int(di["webMSPending_VendorLowLessThan10"])+int(di["eanNonMSPending_VendorCriticalLessThan10"])+int(di["eanNonMSPending_VendorHighLessThan10"])+int(di["eanNonMSPending_VendorAverageLessThan10"])+int(di["eanNonMSPending_VendorLowLessThan10"])+int(di["eanMSPending_VendorCriticalLessThan10"])+int(di["eanMSPending_VendorHighLessThan10"])+int(di["eanMSPending_VendorAverageLessThan10"])+int(di["eanMSPending_VendorLowLessThan10"])+int(di["edsuMSPending_VendorCriticalLessThan10"])+int(di["edsuMSPending_VendorHighLessThan10"])+int(di["edsuMSPending_VendorAverageLessThan10"])+int(di["edsuMSPending_VendorLowLessThan10"])+int(di["edsuNonMSPending_VendorCriticalLessThan10"])+int(di["edsuNonMSPending_VendorHighLessThan10"])+int(di["edsuNonMSPending_VendorAverageLessThan10"])+int(di["edsuNonMSPending_VendorLowLessThan10"]))+'''</td>
            <td>'''+str(int(di["NggMSWork_In_ProgressCriticalLessThan10"])+int(di["NggMSWork_In_ProgressHighLessThan10"])+int(di["NggMSWork_In_ProgressAverageLessThan10"])+int(di["NggMSWork_In_ProgressLowLessThan10"])+int(di["NggNonMSWork_In_ProgressCriticalLessThan10"])+int(di["NggNonMSWork_In_ProgressHighLessThan10"])+int(di["NggNonMSWork_In_ProgressAverageLessThan10"])+int(di["NggNonMSWork_In_ProgressLowLessThan10"])+int(di["webNonMSWork_In_ProgressCriticalLessThan10"])+int(di["webNonMSWork_In_ProgressHighLessThan10"])+int(di["webNonMSWork_In_ProgressAverageLessThan10"])+int(di["webNonMSWork_In_ProgressLowLessThan10"])+int(di["webMSWork_In_ProgressCriticalLessThan10"])+int(di["webMSWork_In_ProgressHighLessThan10"])+int(di["webMSWork_In_ProgressAverageLessThan10"])+int(di["webMSWork_In_ProgressLowLessThan10"])+int(di["eanNonMSWork_In_ProgressCriticalLessThan10"])+int(di["eanNonMSWork_In_ProgressHighLessThan10"])+int(di["eanNonMSWork_In_ProgressAverageLessThan10"])+int(di["eanNonMSWork_In_ProgressLowLessThan10"])+int(di["eanMSWork_In_ProgressCriticalLessThan10"])+int(di["eanMSWork_In_ProgressHighLessThan10"])+int(di["eanMSWork_In_ProgressAverageLessThan10"])+int(di["eanMSWork_In_ProgressLowLessThan10"])+int(di["edsuMSWork_In_ProgressCriticalLessThan10"])+int(di["edsuMSWork_In_ProgressHighLessThan10"])+int(di["edsuMSWork_In_ProgressAverageLessThan10"])+int(di["edsuMSWork_In_ProgressLowLessThan10"])+int(di["edsuNonMSWork_In_ProgressCriticalLessThan10"])+int(di["edsuNonMSWork_In_ProgressHighLessThan10"])+int(di["edsuNonMSWork_In_ProgressAverageLessThan10"])+int(di["edsuNonMSWork_In_ProgressLowLessThan10"]))+'''</td>

        </tr>
        <tr>
            <td class="agenum">>10</td>
            <td>'''+str(int(di["NggMSOpenCriticalGreaterThan10"])+int(di["NggMSOpenHighGreaterThan10"])+int(di["NggMSOpenAverageGreaterThan10"])+int(di["NggMSOpenLowGreaterThan10"])+int(di["NggNonMSOpenCriticalGreaterThan10"])+int(di["NggNonMSOpenHighGreaterThan10"])+int(di["NggNonMSOpenAverageGreaterThan10"])+int(di["NggNonMSOpenLowGreaterThan10"])+int(di["webNonMSOpenCriticalGreaterThan10"])+int(di["webNonMSOpenHighGreaterThan10"])+int(di["webNonMSOpenAverageGreaterThan10"])+int(di["webNonMSOpenLowGreaterThan10"])+int(di["webMSOpenCriticalGreaterThan10"])+int(di["webMSOpenHighGreaterThan10"])+int(di["webMSOpenAverageGreaterThan10"])+int(di["webMSOpenLowGreaterThan10"])+int(di["eanNonMSOpenCriticalGreaterThan10"])+int(di["eanNonMSOpenHighGreaterThan10"])+int(di["eanNonMSOpenAverageGreaterThan10"])+int(di["eanNonMSOpenLowGreaterThan10"])+int(di["eanMSOpenCriticalGreaterThan10"])+int(di["eanMSOpenHighGreaterThan10"])+int(di["eanMSOpenAverageGreaterThan10"])+int(di["eanMSOpenLowGreaterThan10"])+int(di["edsuMSOpenCriticalGreaterThan10"])+int(di["edsuMSOpenHighGreaterThan10"])+int(di["edsuMSOpenAverageGreaterThan10"])+int(di["edsuMSOpenLowGreaterThan10"])+int(di["edsuNonMSOpenCriticalGreaterThan10"])+int(di["edsuNonMSOpenHighGreaterThan10"])+int(di["edsuNonMSOpenAverageGreaterThan10"])+int(di["edsuNonMSOpenLowGreaterThan10"]))+'''</td>
            <td>'''+str(int(di["NggMSPending_ClientCriticalGreaterThan10"])+int(di["NggMSPending_ClientHighGreaterThan10"])+int(di["NggMSPending_ClientAverageGreaterThan10"])+int(di["NggMSPending_ClientLowGreaterThan10"])+int(di["NggNonMSPending_ClientCriticalGreaterThan10"])+int(di["NggNonMSPending_ClientHighGreaterThan10"])+int(di["NggNonMSPending_ClientAverageGreaterThan10"])+int(di["NggNonMSPending_ClientLowGreaterThan10"])+int(di["webNonMSPending_ClientCriticalGreaterThan10"])+int(di["webNonMSPending_ClientHighGreaterThan10"])+int(di["webNonMSPending_ClientAverageGreaterThan10"])+int(di["webNonMSPending_ClientLowGreaterThan10"])+int(di["webMSPending_ClientCriticalGreaterThan10"])+int(di["webMSPending_ClientHighGreaterThan10"])+int(di["webMSPending_ClientAverageGreaterThan10"])+int(di["webMSPending_ClientLowGreaterThan10"])+int(di["eanNonMSPending_ClientCriticalGreaterThan10"])+int(di["eanNonMSPending_ClientHighGreaterThan10"])+int(di["eanNonMSPending_ClientAverageGreaterThan10"])+int(di["eanNonMSPending_ClientLowGreaterThan10"])+int(di["eanMSPending_ClientCriticalGreaterThan10"])+int(di["eanMSPending_ClientHighGreaterThan10"])+int(di["eanMSPending_ClientAverageGreaterThan10"])+int(di["eanMSPending_ClientLowGreaterThan10"])+int(di["edsuMSPending_ClientCriticalGreaterThan10"])+int(di["edsuMSPending_ClientHighGreaterThan10"])+int(di["edsuMSPending_ClientAverageGreaterThan10"])+int(di["edsuMSPending_ClientLowGreaterThan10"])+int(di["edsuNonMSPending_ClientCriticalGreaterThan10"])+int(di["edsuNonMSPending_ClientHighGreaterThan10"])+int(di["edsuNonMSPending_ClientAverageGreaterThan10"])+int(di["edsuNonMSPending_ClientLowGreaterThan10"]))+'''</td>
            <td>'''+str(int(di["NggMSPending_Emergency_ClientCriticalGreaterThan10"])+int(di["NggMSPending_Emergency_ClientHighGreaterThan10"])+int(di["NggMSPending_Emergency_ClientAverageGreaterThan10"])+int(di["NggMSPending_Emergency_ClientLowGreaterThan10"])+int(di["NggNonMSPending_Emergency_ClientCriticalGreaterThan10"])+int(di["NggNonMSPending_Emergency_ClientHighGreaterThan10"])+int(di["NggNonMSPending_Emergency_ClientAverageGreaterThan10"])+int(di["NggNonMSPending_Emergency_ClientLowGreaterThan10"])+int(di["webNonMSPending_Emergency_ClientCriticalGreaterThan10"])+int(di["webNonMSPending_Emergency_ClientHighGreaterThan10"])+int(di["webNonMSPending_Emergency_ClientAverageGreaterThan10"])+int(di["webNonMSPending_Emergency_ClientLowGreaterThan10"])+int(di["webMSPending_Emergency_ClientCriticalGreaterThan10"])+int(di["webMSPending_Emergency_ClientHighGreaterThan10"])+int(di["webMSPending_Emergency_ClientAverageGreaterThan10"])+int(di["webMSPending_Emergency_ClientLowGreaterThan10"])+int(di["eanNonMSPending_Emergency_ClientCriticalGreaterThan10"])+int(di["eanNonMSPending_Emergency_ClientHighGreaterThan10"])+int(di["eanNonMSPending_Emergency_ClientAverageGreaterThan10"])+int(di["eanNonMSPending_Emergency_ClientLowGreaterThan10"])+int(di["eanMSPending_Emergency_ClientCriticalGreaterThan10"])+int(di["eanMSPending_Emergency_ClientHighGreaterThan10"])+int(di["eanMSPending_Emergency_ClientAverageGreaterThan10"])+int(di["eanMSPending_Emergency_ClientLowGreaterThan10"])+int(di["edsuMSPending_Emergency_ClientCriticalGreaterThan10"])+int(di["edsuMSPending_Emergency_ClientHighGreaterThan10"])+int(di["edsuMSPending_Emergency_ClientAverageGreaterThan10"])+int(di["edsuMSPending_Emergency_ClientLowGreaterThan10"])+int(di["edsuNonMSPending_Emergency_ClientCriticalGreaterThan10"])+int(di["edsuNonMSPending_Emergency_ClientHighGreaterThan10"])+int(di["edsuNonMSPending_Emergency_ClientAverageGreaterThan10"])+int(di["edsuNonMSPending_Emergency_ClientLowGreaterThan10"]))+'''</td>
            <td>'''+str(int(di["NggMSPending_VendorCriticalGreaterThan10"])+int(di["NggMSPending_VendorHighGreaterThan10"])+int(di["NggMSPending_VendorAverageGreaterThan10"])+int(di["NggMSPending_VendorLowGreaterThan10"])+int(di["NggNonMSPending_VendorCriticalGreaterThan10"])+int(di["NggNonMSPending_VendorHighGreaterThan10"])+int(di["NggNonMSPending_VendorAverageGreaterThan10"])+int(di["NggNonMSPending_VendorLowGreaterThan10"])+int(di["webNonMSPending_VendorCriticalGreaterThan10"])+int(di["webNonMSPending_VendorHighGreaterThan10"])+int(di["webNonMSPending_VendorAverageGreaterThan10"])+int(di["webNonMSPending_VendorLowGreaterThan10"])+int(di["webMSPending_VendorCriticalGreaterThan10"])+int(di["webMSPending_VendorHighGreaterThan10"])+int(di["webMSPending_VendorAverageGreaterThan10"])+int(di["webMSPending_VendorLowGreaterThan10"])+int(di["eanNonMSPending_VendorCriticalGreaterThan10"])+int(di["eanNonMSPending_VendorHighGreaterThan10"])+int(di["eanNonMSPending_VendorAverageGreaterThan10"])+int(di["eanNonMSPending_VendorLowGreaterThan10"])+int(di["eanMSPending_VendorCriticalGreaterThan10"])+int(di["eanMSPending_VendorHighGreaterThan10"])+int(di["eanMSPending_VendorAverageGreaterThan10"])+int(di["eanMSPending_VendorLowGreaterThan10"])+int(di["edsuMSPending_VendorCriticalGreaterThan10"])+int(di["edsuMSPending_VendorHighGreaterThan10"])+int(di["edsuMSPending_VendorAverageGreaterThan10"])+int(di["edsuMSPending_VendorLowGreaterThan10"])+int(di["edsuNonMSPending_VendorCriticalGreaterThan10"])+int(di["edsuNonMSPending_VendorHighGreaterThan10"])+int(di["edsuNonMSPending_VendorAverageGreaterThan10"])+int(di["edsuNonMSPending_VendorLowGreaterThan10"]))+'''</td>
            <td>'''+str(int(di["NggMSWork_In_ProgressCriticalGreaterThan10"])+int(di["NggMSWork_In_ProgressHighGreaterThan10"])+int(di["NggMSWork_In_ProgressAverageGreaterThan10"])+int(di["NggMSWork_In_ProgressLowGreaterThan10"])+int(di["NggNonMSWork_In_ProgressCriticalGreaterThan10"])+int(di["NggNonMSWork_In_ProgressHighGreaterThan10"])+int(di["NggNonMSWork_In_ProgressAverageGreaterThan10"])+int(di["NggNonMSWork_In_ProgressLowGreaterThan10"])+int(di["webNonMSWork_In_ProgressCriticalGreaterThan10"])+int(di["webNonMSWork_In_ProgressHighGreaterThan10"])+int(di["webNonMSWork_In_ProgressAverageGreaterThan10"])+int(di["webNonMSWork_In_ProgressLowGreaterThan10"])+int(di["webMSWork_In_ProgressCriticalGreaterThan10"])+int(di["webMSWork_In_ProgressHighGreaterThan10"])+int(di["webMSWork_In_ProgressAverageGreaterThan10"])+int(di["webMSWork_In_ProgressLowGreaterThan10"])+int(di["eanNonMSWork_In_ProgressCriticalGreaterThan10"])+int(di["eanNonMSWork_In_ProgressHighGreaterThan10"])+int(di["eanNonMSWork_In_ProgressAverageGreaterThan10"])+int(di["eanNonMSWork_In_ProgressLowGreaterThan10"])+int(di["eanMSWork_In_ProgressCriticalGreaterThan10"])+int(di["eanMSWork_In_ProgressHighGreaterThan10"])+int(di["eanMSWork_In_ProgressAverageGreaterThan10"])+int(di["eanMSWork_In_ProgressLowGreaterThan10"])+int(di["edsuMSWork_In_ProgressCriticalGreaterThan10"])+int(di["edsuMSWork_In_ProgressHighGreaterThan10"])+int(di["edsuMSWork_In_ProgressAverageGreaterThan10"])+int(di["edsuMSWork_In_ProgressLowGreaterThan10"])+int(di["edsuNonMSWork_In_ProgressCriticalGreaterThan10"])+int(di["edsuNonMSWork_In_ProgressHighGreaterThan10"])+int(di["edsuNonMSWork_In_ProgressAverageGreaterThan10"])+int(di["edsuNonMSWork_In_ProgressLowGreaterThan10"]))+'''</td>

        </tr>
        <tr>
            <td class="agenum">>50</td>
            <td>'''+str(int(di["NggMSOpenCriticalGreaterThan50"])+int(di["NggMSOpenHighGreaterThan50"])+int(di["NggMSOpenAverageGreaterThan50"])+int(di["NggMSOpenLowGreaterThan50"])+int(di["NggNonMSOpenCriticalGreaterThan50"])+int(di["NggNonMSOpenHighGreaterThan50"])+int(di["NggNonMSOpenAverageGreaterThan50"])+int(di["NggNonMSOpenLowGreaterThan50"])+int(di["webNonMSOpenCriticalGreaterThan50"])+int(di["webNonMSOpenHighGreaterThan50"])+int(di["webNonMSOpenAverageGreaterThan50"])+int(di["webNonMSOpenLowGreaterThan50"])+int(di["webMSOpenCriticalGreaterThan50"])+int(di["webMSOpenHighGreaterThan50"])+int(di["webMSOpenAverageGreaterThan50"])+int(di["webMSOpenLowGreaterThan50"])+int(di["eanNonMSOpenCriticalGreaterThan50"])+int(di["eanNonMSOpenHighGreaterThan50"])+int(di["eanNonMSOpenAverageGreaterThan50"])+int(di["eanNonMSOpenLowGreaterThan50"])+int(di["eanMSOpenCriticalGreaterThan50"])+int(di["eanMSOpenHighGreaterThan50"])+int(di["eanMSOpenAverageGreaterThan50"])+int(di["eanMSOpenLowGreaterThan50"])+int(di["edsuMSOpenCriticalGreaterThan50"])+int(di["edsuMSOpenHighGreaterThan50"])+int(di["edsuMSOpenAverageGreaterThan50"])+int(di["edsuMSOpenLowGreaterThan50"])+int(di["edsuNonMSOpenCriticalGreaterThan50"])+int(di["edsuNonMSOpenHighGreaterThan50"])+int(di["edsuNonMSOpenAverageGreaterThan50"])+int(di["edsuNonMSOpenLowGreaterThan50"]))+'''</td>
            <td>'''+str(int(di["NggMSPending_ClientCriticalGreaterThan50"])+int(di["NggMSPending_ClientHighGreaterThan50"])+int(di["NggMSPending_ClientAverageGreaterThan50"])+int(di["NggMSPending_ClientLowGreaterThan50"])+int(di["NggNonMSPending_ClientCriticalGreaterThan50"])+int(di["NggNonMSPending_ClientHighGreaterThan50"])+int(di["NggNonMSPending_ClientAverageGreaterThan50"])+int(di["NggNonMSPending_ClientLowGreaterThan50"])+int(di["webNonMSPending_ClientCriticalGreaterThan50"])+int(di["webNonMSPending_ClientHighGreaterThan50"])+int(di["webNonMSPending_ClientAverageGreaterThan50"])+int(di["webNonMSPending_ClientLowGreaterThan50"])+int(di["webMSPending_ClientCriticalGreaterThan50"])+int(di["webMSPending_ClientHighGreaterThan50"])+int(di["webMSPending_ClientAverageGreaterThan50"])+int(di["webMSPending_ClientLowGreaterThan50"])+int(di["eanNonMSPending_ClientCriticalGreaterThan50"])+int(di["eanNonMSPending_ClientHighGreaterThan50"])+int(di["eanNonMSPending_ClientAverageGreaterThan50"])+int(di["eanNonMSPending_ClientLowGreaterThan50"])+int(di["eanMSPending_ClientCriticalGreaterThan50"])+int(di["eanMSPending_ClientHighGreaterThan50"])+int(di["eanMSPending_ClientAverageGreaterThan50"])+int(di["eanMSPending_ClientLowGreaterThan50"])+int(di["edsuMSPending_ClientCriticalGreaterThan50"])+int(di["edsuMSPending_ClientHighGreaterThan50"])+int(di["edsuMSPending_ClientAverageGreaterThan50"])+int(di["edsuMSPending_ClientLowGreaterThan50"])+int(di["edsuNonMSPending_ClientCriticalGreaterThan50"])+int(di["edsuNonMSPending_ClientHighGreaterThan50"])+int(di["edsuNonMSPending_ClientAverageGreaterThan50"])+int(di["edsuNonMSPending_ClientLowGreaterThan50"]))+'''</td>
            <td>'''+str(int(di["NggMSPending_Emergency_ClientCriticalGreaterThan50"])+int(di["NggMSPending_Emergency_ClientHighGreaterThan50"])+int(di["NggMSPending_Emergency_ClientAverageGreaterThan50"])+int(di["NggMSPending_Emergency_ClientLowGreaterThan50"])+int(di["NggNonMSPending_Emergency_ClientCriticalGreaterThan50"])+int(di["NggNonMSPending_Emergency_ClientHighGreaterThan50"])+int(di["NggNonMSPending_Emergency_ClientAverageGreaterThan50"])+int(di["NggNonMSPending_Emergency_ClientLowGreaterThan50"])+int(di["webNonMSPending_Emergency_ClientCriticalGreaterThan50"])+int(di["webNonMSPending_Emergency_ClientHighGreaterThan50"])+int(di["webNonMSPending_Emergency_ClientAverageGreaterThan50"])+int(di["webNonMSPending_Emergency_ClientLowGreaterThan50"])+int(di["webMSPending_Emergency_ClientCriticalGreaterThan50"])+int(di["webMSPending_Emergency_ClientHighGreaterThan50"])+int(di["webMSPending_Emergency_ClientAverageGreaterThan50"])+int(di["webMSPending_Emergency_ClientLowGreaterThan50"])+int(di["eanNonMSPending_Emergency_ClientCriticalGreaterThan50"])+int(di["eanNonMSPending_Emergency_ClientHighGreaterThan50"])+int(di["eanNonMSPending_Emergency_ClientAverageGreaterThan50"])+int(di["eanNonMSPending_Emergency_ClientLowGreaterThan50"])+int(di["eanMSPending_Emergency_ClientCriticalGreaterThan50"])+int(di["eanMSPending_Emergency_ClientHighGreaterThan50"])+int(di["eanMSPending_Emergency_ClientAverageGreaterThan50"])+int(di["eanMSPending_Emergency_ClientLowGreaterThan50"])+int(di["edsuMSPending_Emergency_ClientCriticalGreaterThan50"])+int(di["edsuMSPending_Emergency_ClientHighGreaterThan50"])+int(di["edsuMSPending_Emergency_ClientAverageGreaterThan50"])+int(di["edsuMSPending_Emergency_ClientLowGreaterThan50"])+int(di["edsuNonMSPending_Emergency_ClientCriticalGreaterThan50"])+int(di["edsuNonMSPending_Emergency_ClientHighGreaterThan50"])+int(di["edsuNonMSPending_Emergency_ClientAverageGreaterThan50"])+int(di["edsuNonMSPending_Emergency_ClientLowGreaterThan50"]))+'''</td>
            <td>'''+str(int(di["NggMSPending_VendorCriticalGreaterThan50"])+int(di["NggMSPending_VendorHighGreaterThan50"])+int(di["NggMSPending_VendorAverageGreaterThan50"])+int(di["NggMSPending_VendorLowGreaterThan50"])+int(di["NggNonMSPending_VendorCriticalGreaterThan50"])+int(di["NggNonMSPending_VendorHighGreaterThan50"])+int(di["NggNonMSPending_VendorAverageGreaterThan50"])+int(di["NggNonMSPending_VendorLowGreaterThan50"])+int(di["webNonMSPending_VendorCriticalGreaterThan50"])+int(di["webNonMSPending_VendorHighGreaterThan50"])+int(di["webNonMSPending_VendorAverageGreaterThan50"])+int(di["webNonMSPending_VendorLowGreaterThan50"])+int(di["webMSPending_VendorCriticalGreaterThan50"])+int(di["webMSPending_VendorHighGreaterThan50"])+int(di["webMSPending_VendorAverageGreaterThan50"])+int(di["webMSPending_VendorLowGreaterThan50"])+int(di["eanNonMSPending_VendorCriticalGreaterThan50"])+int(di["eanNonMSPending_VendorHighGreaterThan50"])+int(di["eanNonMSPending_VendorAverageGreaterThan50"])+int(di["eanNonMSPending_VendorLowGreaterThan50"])+int(di["eanMSPending_VendorCriticalGreaterThan50"])+int(di["eanMSPending_VendorHighGreaterThan50"])+int(di["eanMSPending_VendorAverageGreaterThan50"])+int(di["eanMSPending_VendorLowGreaterThan50"])+int(di["edsuMSPending_VendorCriticalGreaterThan50"])+int(di["edsuMSPending_VendorHighGreaterThan50"])+int(di["edsuMSPending_VendorAverageGreaterThan50"])+int(di["edsuMSPending_VendorLowGreaterThan50"])+int(di["edsuNonMSPending_VendorCriticalGreaterThan50"])+int(di["edsuNonMSPending_VendorHighGreaterThan50"])+int(di["edsuNonMSPending_VendorAverageGreaterThan50"])+int(di["edsuNonMSPending_VendorLowGreaterThan50"]))+'''</td>
            <td>'''+str(int(di["NggMSWork_In_ProgressCriticalGreaterThan50"])+int(di["NggMSWork_In_ProgressHighGreaterThan50"])+int(di["NggMSWork_In_ProgressAverageGreaterThan50"])+int(di["NggMSWork_In_ProgressLowGreaterThan50"])+int(di["NggNonMSWork_In_ProgressCriticalGreaterThan50"])+int(di["NggNonMSWork_In_ProgressHighGreaterThan50"])+int(di["NggNonMSWork_In_ProgressAverageGreaterThan50"])+int(di["NggNonMSWork_In_ProgressLowGreaterThan50"])+int(di["webNonMSWork_In_ProgressCriticalGreaterThan50"])+int(di["webNonMSWork_In_ProgressHighGreaterThan50"])+int(di["webNonMSWork_In_ProgressAverageGreaterThan50"])+int(di["webNonMSWork_In_ProgressLowGreaterThan50"])+int(di["webMSWork_In_ProgressCriticalGreaterThan50"])+int(di["webMSWork_In_ProgressHighGreaterThan50"])+int(di["webMSWork_In_ProgressAverageGreaterThan50"])+int(di["webMSWork_In_ProgressLowGreaterThan50"])+int(di["eanNonMSWork_In_ProgressCriticalGreaterThan50"])+int(di["eanNonMSWork_In_ProgressHighGreaterThan50"])+int(di["eanNonMSWork_In_ProgressAverageGreaterThan50"])+int(di["eanNonMSWork_In_ProgressLowGreaterThan50"])+int(di["eanMSWork_In_ProgressCriticalGreaterThan50"])+int(di["eanMSWork_In_ProgressHighGreaterThan50"])+int(di["eanMSWork_In_ProgressAverageGreaterThan50"])+int(di["eanMSWork_In_ProgressLowGreaterThan50"])+int(di["edsuMSWork_In_ProgressCriticalGreaterThan50"])+int(di["edsuMSWork_In_ProgressHighGreaterThan50"])+int(di["edsuMSWork_In_ProgressAverageGreaterThan50"])+int(di["edsuMSWork_In_ProgressLowGreaterThan50"])+int(di["edsuNonMSWork_In_ProgressCriticalGreaterThan50"])+int(di["edsuNonMSWork_In_ProgressHighGreaterThan50"])+int(di["edsuNonMSWork_In_ProgressAverageGreaterThan50"])+int(di["edsuNonMSWork_In_ProgressLowGreaterThan50"]))+'''</td>

        </tr>
        <tr>
            <td class="agenum">>100</td>
            <td>'''+str(int(di["NggMSOpenCriticalGreaterThan100"])+int(di["NggMSOpenHighGreaterThan100"])+int(di["NggMSOpenAverageGreaterThan100"])+int(di["NggMSOpenLowGreaterThan100"])+int(di["NggNonMSOpenCriticalGreaterThan100"])+int(di["NggNonMSOpenHighGreaterThan100"])+int(di["NggNonMSOpenAverageGreaterThan100"])+int(di["NggNonMSOpenLowGreaterThan100"])+int(di["webNonMSOpenCriticalGreaterThan100"])+int(di["webNonMSOpenHighGreaterThan100"])+int(di["webNonMSOpenAverageGreaterThan100"])+int(di["webNonMSOpenLowGreaterThan100"])+int(di["webMSOpenCriticalGreaterThan100"])+int(di["webMSOpenHighGreaterThan100"])+int(di["webMSOpenAverageGreaterThan100"])+int(di["webMSOpenLowGreaterThan100"])+int(di["eanNonMSOpenCriticalGreaterThan100"])+int(di["eanNonMSOpenHighGreaterThan100"])+int(di["eanNonMSOpenAverageGreaterThan100"])+int(di["eanNonMSOpenLowGreaterThan100"])+int(di["eanMSOpenCriticalGreaterThan100"])+int(di["eanMSOpenHighGreaterThan100"])+int(di["eanMSOpenAverageGreaterThan100"])+int(di["eanMSOpenLowGreaterThan100"])+int(di["edsuMSOpenCriticalGreaterThan100"])+int(di["edsuMSOpenHighGreaterThan100"])+int(di["edsuMSOpenAverageGreaterThan100"])+int(di["edsuMSOpenLowGreaterThan100"])+int(di["edsuNonMSOpenCriticalGreaterThan100"])+int(di["edsuNonMSOpenHighGreaterThan100"])+int(di["edsuNonMSOpenAverageGreaterThan100"])+int(di["edsuNonMSOpenLowGreaterThan100"]))+'''</td>
            <td>'''+str(int(di["NggMSPending_ClientCriticalGreaterThan100"])+int(di["NggMSPending_ClientHighGreaterThan100"])+int(di["NggMSPending_ClientAverageGreaterThan100"])+int(di["NggMSPending_ClientLowGreaterThan100"])+int(di["NggNonMSPending_ClientCriticalGreaterThan100"])+int(di["NggNonMSPending_ClientHighGreaterThan100"])+int(di["NggNonMSPending_ClientAverageGreaterThan100"])+int(di["NggNonMSPending_ClientLowGreaterThan100"])+int(di["webNonMSPending_ClientCriticalGreaterThan100"])+int(di["webNonMSPending_ClientHighGreaterThan100"])+int(di["webNonMSPending_ClientAverageGreaterThan100"])+int(di["webNonMSPending_ClientLowGreaterThan100"])+int(di["webMSPending_ClientCriticalGreaterThan100"])+int(di["webMSPending_ClientHighGreaterThan100"])+int(di["webMSPending_ClientAverageGreaterThan100"])+int(di["webMSPending_ClientLowGreaterThan100"])+int(di["eanNonMSPending_ClientCriticalGreaterThan100"])+int(di["eanNonMSPending_ClientHighGreaterThan100"])+int(di["eanNonMSPending_ClientAverageGreaterThan100"])+int(di["eanNonMSPending_ClientLowGreaterThan100"])+int(di["eanMSPending_ClientCriticalGreaterThan100"])+int(di["eanMSPending_ClientHighGreaterThan100"])+int(di["eanMSPending_ClientAverageGreaterThan100"])+int(di["eanMSPending_ClientLowGreaterThan100"])+int(di["edsuMSPending_ClientCriticalGreaterThan100"])+int(di["edsuMSPending_ClientHighGreaterThan100"])+int(di["edsuMSPending_ClientAverageGreaterThan100"])+int(di["edsuMSPending_ClientLowGreaterThan100"])+int(di["edsuNonMSPending_ClientCriticalGreaterThan100"])+int(di["edsuNonMSPending_ClientHighGreaterThan100"])+int(di["edsuNonMSPending_ClientAverageGreaterThan100"])+int(di["edsuNonMSPending_ClientLowGreaterThan100"]))+'''</td>
            <td>'''+str(int(di["NggMSPending_Emergency_ClientCriticalGreaterThan100"])+int(di["NggMSPending_Emergency_ClientHighGreaterThan100"])+int(di["NggMSPending_Emergency_ClientAverageGreaterThan100"])+int(di["NggMSPending_Emergency_ClientLowGreaterThan100"])+int(di["NggNonMSPending_Emergency_ClientCriticalGreaterThan100"])+int(di["NggNonMSPending_Emergency_ClientHighGreaterThan100"])+int(di["NggNonMSPending_Emergency_ClientAverageGreaterThan100"])+int(di["NggNonMSPending_Emergency_ClientLowGreaterThan100"])+int(di["webNonMSPending_Emergency_ClientCriticalGreaterThan100"])+int(di["webNonMSPending_Emergency_ClientHighGreaterThan100"])+int(di["webNonMSPending_Emergency_ClientAverageGreaterThan100"])+int(di["webNonMSPending_Emergency_ClientLowGreaterThan100"])+int(di["webMSPending_Emergency_ClientCriticalGreaterThan100"])+int(di["webMSPending_Emergency_ClientHighGreaterThan100"])+int(di["webMSPending_Emergency_ClientAverageGreaterThan100"])+int(di["webMSPending_Emergency_ClientLowGreaterThan100"])+int(di["eanNonMSPending_Emergency_ClientCriticalGreaterThan100"])+int(di["eanNonMSPending_Emergency_ClientHighGreaterThan100"])+int(di["eanNonMSPending_Emergency_ClientAverageGreaterThan100"])+int(di["eanNonMSPending_Emergency_ClientLowGreaterThan100"])+int(di["eanMSPending_Emergency_ClientCriticalGreaterThan100"])+int(di["eanMSPending_Emergency_ClientHighGreaterThan100"])+int(di["eanMSPending_Emergency_ClientAverageGreaterThan100"])+int(di["eanMSPending_Emergency_ClientLowGreaterThan100"])+int(di["edsuMSPending_Emergency_ClientCriticalGreaterThan100"])+int(di["edsuMSPending_Emergency_ClientHighGreaterThan100"])+int(di["edsuMSPending_Emergency_ClientAverageGreaterThan100"])+int(di["edsuMSPending_Emergency_ClientLowGreaterThan100"])+int(di["edsuNonMSPending_Emergency_ClientCriticalGreaterThan100"])+int(di["edsuNonMSPending_Emergency_ClientHighGreaterThan100"])+int(di["edsuNonMSPending_Emergency_ClientAverageGreaterThan100"])+int(di["edsuNonMSPending_Emergency_ClientLowGreaterThan100"]))+'''</td>
            <td>'''+str(int(di["NggMSPending_VendorCriticalGreaterThan100"])+int(di["NggMSPending_VendorHighGreaterThan100"])+int(di["NggMSPending_VendorAverageGreaterThan100"])+int(di["NggMSPending_VendorLowGreaterThan100"])+int(di["NggNonMSPending_VendorCriticalGreaterThan100"])+int(di["NggNonMSPending_VendorHighGreaterThan100"])+int(di["NggNonMSPending_VendorAverageGreaterThan100"])+int(di["NggNonMSPending_VendorLowGreaterThan100"])+int(di["webNonMSPending_VendorCriticalGreaterThan100"])+int(di["webNonMSPending_VendorHighGreaterThan100"])+int(di["webNonMSPending_VendorAverageGreaterThan100"])+int(di["webNonMSPending_VendorLowGreaterThan100"])+int(di["webMSPending_VendorCriticalGreaterThan100"])+int(di["webMSPending_VendorHighGreaterThan100"])+int(di["webMSPending_VendorAverageGreaterThan100"])+int(di["webMSPending_VendorLowGreaterThan100"])+int(di["eanNonMSPending_VendorCriticalGreaterThan100"])+int(di["eanNonMSPending_VendorHighGreaterThan100"])+int(di["eanNonMSPending_VendorAverageGreaterThan100"])+int(di["eanNonMSPending_VendorLowGreaterThan100"])+int(di["eanMSPending_VendorCriticalGreaterThan100"])+int(di["eanMSPending_VendorHighGreaterThan100"])+int(di["eanMSPending_VendorAverageGreaterThan100"])+int(di["eanMSPending_VendorLowGreaterThan100"])+int(di["edsuMSPending_VendorCriticalGreaterThan100"])+int(di["edsuMSPending_VendorHighGreaterThan100"])+int(di["edsuMSPending_VendorAverageGreaterThan100"])+int(di["edsuMSPending_VendorLowGreaterThan100"])+int(di["edsuNonMSPending_VendorCriticalGreaterThan100"])+int(di["edsuNonMSPending_VendorHighGreaterThan100"])+int(di["edsuNonMSPending_VendorAverageGreaterThan100"])+int(di["edsuNonMSPending_VendorLowGreaterThan100"]))+'''</td>
            <td>'''+str(int(di["NggMSWork_In_ProgressCriticalGreaterThan100"])+int(di["NggMSWork_In_ProgressHighGreaterThan100"])+int(di["NggMSWork_In_ProgressAverageGreaterThan100"])+int(di["NggMSWork_In_ProgressLowGreaterThan100"])+int(di["NggNonMSWork_In_ProgressCriticalGreaterThan100"])+int(di["NggNonMSWork_In_ProgressHighGreaterThan100"])+int(di["NggNonMSWork_In_ProgressAverageGreaterThan100"])+int(di["NggNonMSWork_In_ProgressLowGreaterThan100"])+int(di["webNonMSWork_In_ProgressCriticalGreaterThan100"])+int(di["webNonMSWork_In_ProgressHighGreaterThan100"])+int(di["webNonMSWork_In_ProgressAverageGreaterThan100"])+int(di["webNonMSWork_In_ProgressLowGreaterThan100"])+int(di["webMSWork_In_ProgressCriticalGreaterThan100"])+int(di["webMSWork_In_ProgressHighGreaterThan100"])+int(di["webMSWork_In_ProgressAverageGreaterThan100"])+int(di["webMSWork_In_ProgressLowGreaterThan100"])+int(di["eanNonMSWork_In_ProgressCriticalGreaterThan100"])+int(di["eanNonMSWork_In_ProgressHighGreaterThan100"])+int(di["eanNonMSWork_In_ProgressAverageGreaterThan100"])+int(di["eanNonMSWork_In_ProgressLowGreaterThan100"])+int(di["eanMSWork_In_ProgressCriticalGreaterThan100"])+int(di["eanMSWork_In_ProgressHighGreaterThan100"])+int(di["eanMSWork_In_ProgressAverageGreaterThan100"])+int(di["eanMSWork_In_ProgressLowGreaterThan100"])+int(di["edsuMSWork_In_ProgressCriticalGreaterThan100"])+int(di["edsuMSWork_In_ProgressHighGreaterThan100"])+int(di["edsuMSWork_In_ProgressAverageGreaterThan100"])+int(di["edsuMSWork_In_ProgressLowGreaterThan100"])+int(di["edsuNonMSWork_In_ProgressCriticalGreaterThan100"])+int(di["edsuNonMSWork_In_ProgressHighGreaterThan100"])+int(di["edsuNonMSWork_In_ProgressAverageGreaterThan100"])+int(di["edsuNonMSWork_In_ProgressLowGreaterThan100"]))+'''</td>

        </tr>
        </table>
    </div>
    <br>
    </br>
    <div>
    <table  border="2pxsingleblack">
        <tr class="supportgroup" >
            <td colspan="14" style="text-align:center;" >US_NGG-APPL-Support</td>
        </tr>
        <tr bgcolor="#FED8B1">
            <td colspan="7" style="text-align:center;">MS Support Group</td>
            <td colspan="7" style="text-align:center;">Non MS Support Group</td>
        </tr>
        <tr>
            <td rowspan="5" bgcolor="#ADD8E6">Open</td>
            <td bgcolor="#ffcccb">Aging</td>
            <td  bgcolor="#D3D3D3">Critical</td>
            <td  bgcolor="#D3D3D3">High</td>
            <td  bgcolor="#D3D3D3">Average</td>
            <td  bgcolor="#D3D3D3">Low</td>
            <td  bgcolor="#D3D3D3">Total</td>
            <td rowspan="5" bgcolor="#ADD8E6">Open</td>
            <td bgcolor="#ffcccb">Aging</td>
            <td  bgcolor="#D3D3D3">Critical</td>
            <td  bgcolor="#D3D3D3">High</td>
            <td  bgcolor="#D3D3D3">Average</td>
            <td  bgcolor="#D3D3D3">Low</td>
            <td  bgcolor="#D3D3D3">Total</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["NggMSOpenCriticalLessThan10"]+'''</td>
            <td >'''+di["NggMSOpenHighLessThan10"]+'''</td>
            <td>'''+di["NggMSOpenAverageLessThan10"]+'''</td>
            <td>'''+di["NggMSOpenLowLessThan10"]+'''</td>
            <td>'''+str(int(di["NggMSOpenCriticalLessThan10"])+int(di["NggMSOpenHighLessThan10"])+int(di["NggMSOpenAverageLessThan10"])+int(di["NggMSOpenLowLessThan10"]))+'''</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["NggNonMSOpenCriticalLessThan10"]+'''</td>
            <td >'''+di["NggNonMSOpenHighLessThan10"]+'''</td>
            <td>'''+di["NggNonMSOpenAverageLessThan10"]+'''</td>
            <td>'''+di["NggNonMSOpenLowLessThan10"]+'''</td>
            <td>'''+str(int(di["NggNonMSOpenCriticalLessThan10"])+int(di["NggNonMSOpenHighLessThan10"])+int(di["NggNonMSOpenAverageLessThan10"])+int(di["NggNonMSOpenLowLessThan10"]))+'''</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["NggMSOpenCriticalGreaterThan10"]+'''</td>
            <td >'''+di["NggMSOpenHighGreaterThan10"]+'''</td>
            <td>'''+di["NggMSOpenAverageGreaterThan10"]+'''</td>
            <td>'''+di["NggMSOpenLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["NggMSOpenCriticalGreaterThan10"])+int(di["NggMSOpenHighGreaterThan10"])+int(di["NggMSOpenAverageGreaterThan10"])+int(di["NggMSOpenLowGreaterThan10"]))+'''</td>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["NggNonMSOpenCriticalGreaterThan10"]+'''</td>
            <td >'''+di["NggNonMSOpenHighGreaterThan10"]+'''</td>
            <td>'''+di["NggNonMSOpenAverageGreaterThan10"]+'''</td>
            <td>'''+di["NggNonMSOpenLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["NggNonMSOpenCriticalGreaterThan10"])+int(di["NggNonMSOpenHighGreaterThan10"])+int(di["NggNonMSOpenAverageGreaterThan10"])+int(di["NggNonMSOpenLowGreaterThan10"]))+'''</td>
        </tr>

        <tr>
            <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["NggMSOpenCriticalGreaterThan50"]+'''</td>
            <td >'''+di["NggMSOpenHighGreaterThan50"]+'''</td>
            <td >'''+di["NggMSOpenAverageGreaterThan50"]+'''</td>
            <td >'''+di["NggMSOpenLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["NggMSOpenCriticalGreaterThan50"])+int(di["NggMSOpenHighGreaterThan50"])+int(di["NggMSOpenAverageGreaterThan50"])+int(di["NggMSOpenLowGreaterThan50"]))+'''</td>
             <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["NggNonMSOpenCriticalGreaterThan50"]+'''</td>
            <td >'''+di["NggNonMSOpenHighGreaterThan50"]+'''</td>
            <td >'''+di["NggNonMSOpenAverageGreaterThan50"]+'''</td>
            <td >'''+di["NggNonMSOpenLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["NggNonMSOpenCriticalGreaterThan50"])+int(di["NggNonMSOpenHighGreaterThan50"])+int(di["NggNonMSOpenAverageGreaterThan50"])+int(di["NggNonMSOpenLowGreaterThan50"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["NggMSOpenCriticalGreaterThan100"]+'''</td>
            <td >'''+di["NggMSOpenHighGreaterThan100"]+'''</td>
            <td >'''+di["NggMSOpenAverageGreaterThan100"]+'''</td>
            <td >'''+di["NggMSOpenLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["NggMSOpenCriticalGreaterThan100"])+int(di["NggMSOpenHighGreaterThan100"])+int(di["NggMSOpenAverageGreaterThan100"])+int(di["NggMSOpenLowGreaterThan100"]))+'''</td>
             <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["NggNonMSOpenCriticalGreaterThan100"]+'''</td>
            <td >'''+di["NggNonMSOpenHighGreaterThan100"]+'''</td>
            <td >'''+di["NggNonMSOpenAverageGreaterThan100"]+'''</td>
            <td >'''+di["NggNonMSOpenLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["NggNonMSOpenCriticalGreaterThan100"])+int(di["NggNonMSOpenHighGreaterThan100"])+int(di["NggNonMSOpenAverageGreaterThan100"])+int(di["NggNonMSOpenLowGreaterThan100"]))+'''</td>
        </tr>

        <tr>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Client</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["NggMSPending_ClientCriticalLessThan10"]+'''</td>
            <td >'''+di["NggMSPending_ClientHighLessThan10"]+'''</td>
            <td>'''+di["NggMSPending_ClientAverageLessThan10"]+'''</td>
            <td>'''+di["NggMSPending_ClientLowLessThan10"]+'''</td>
            <td>'''+str(int(di["NggMSPending_ClientCriticalLessThan10"])+int(di["NggMSPending_ClientHighLessThan10"])+int(di["NggMSPending_ClientAverageLessThan10"])+int(di["NggMSPending_ClientLowLessThan10"]))+'''</td>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Client</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["NggNonMSPending_ClientCriticalLessThan10"]+'''</td>
            <td >'''+di["NggNonMSPending_ClientHighLessThan10"]+'''</td>
            <td>'''+di["NggNonMSPending_ClientAverageLessThan10"]+'''</td>
            <td>'''+di["NggNonMSPending_ClientLowLessThan10"]+'''</td>
            <td>'''+str(int(di["NggNonMSPending_ClientCriticalLessThan10"])+int(di["NggNonMSPending_ClientHighLessThan10"])+int(di["NggNonMSPending_ClientAverageLessThan10"])+int(di["NggNonMSPending_ClientLowLessThan10"]))+'''</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["NggMSPending_ClientCriticalGreaterThan10"]+'''</td>
            <td >'''+di["NggMSPending_ClientHighGreaterThan10"]+'''</td>
            <td>'''+di["NggMSPending_ClientAverageGreaterThan10"]+'''</td>
            <td>'''+di["NggMSPending_ClientLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["NggMSPending_ClientCriticalGreaterThan10"])+int(di["NggMSPending_ClientHighGreaterThan10"])+int(di["NggMSPending_ClientAverageGreaterThan10"])+int(di["NggMSPending_ClientLowGreaterThan10"]))+'''</td>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["NggNonMSPending_ClientCriticalGreaterThan10"]+'''</td>
            <td >'''+di["NggNonMSPending_ClientHighGreaterThan10"]+'''</td>
            <td>'''+di["NggNonMSPending_ClientAverageGreaterThan10"]+'''</td>
            <td>'''+di["NggNonMSPending_ClientLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["NggNonMSPending_ClientCriticalGreaterThan10"])+int(di["NggNonMSPending_ClientHighGreaterThan10"])+int(di["NggNonMSPending_ClientAverageGreaterThan10"])+int(di["NggNonMSPending_ClientLowGreaterThan10"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["NggMSPending_ClientCriticalGreaterThan50"]+'''</td>
            <td >'''+di["NggMSPending_ClientHighGreaterThan50"]+'''</td>
            <td >'''+di["NggMSPending_ClientAverageGreaterThan50"]+'''</td>
            <td >'''+di["NggMSPending_ClientLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["NggMSPending_ClientCriticalGreaterThan50"])+int(di["NggMSPending_ClientHighGreaterThan50"])+int(di["NggMSPending_ClientAverageGreaterThan50"])+int(di["NggMSPending_ClientLowGreaterThan50"]))+'''</td>
             <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["NggNonMSPending_ClientCriticalGreaterThan50"]+'''</td>
            <td >'''+di["NggNonMSPending_ClientHighGreaterThan50"]+'''</td>
            <td >'''+di["NggNonMSPending_ClientAverageGreaterThan50"]+'''</td>
            <td >'''+di["NggNonMSPending_ClientLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["NggNonMSPending_ClientCriticalGreaterThan50"])+int(di["NggNonMSPending_ClientHighGreaterThan50"])+int(di["NggNonMSPending_ClientAverageGreaterThan50"])+int(di["NggNonMSPending_ClientLowGreaterThan50"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["NggMSPending_ClientCriticalGreaterThan100"]+'''</td>
            <td >'''+di["NggMSPending_ClientHighGreaterThan100"]+'''</td>
            <td >'''+di["NggMSPending_ClientAverageGreaterThan100"]+'''</td>
            <td >'''+di["NggMSPending_ClientLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["NggMSPending_ClientCriticalGreaterThan100"])+int(di["NggMSPending_ClientHighGreaterThan100"])+int(di["NggMSPending_ClientAverageGreaterThan100"])+int(di["NggMSPending_ClientLowGreaterThan100"]))+'''</td>
             <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["NggNonMSPending_ClientCriticalGreaterThan100"]+'''</td>
            <td >'''+di["NggNonMSPending_ClientHighGreaterThan100"]+'''</td>
            <td >'''+di["NggNonMSPending_ClientAverageGreaterThan100"]+'''</td>
            <td >'''+di["NggNonMSPending_ClientLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["NggNonMSPending_ClientCriticalGreaterThan100"])+int(di["NggNonMSPending_ClientHighGreaterThan100"])+int(di["NggNonMSPending_ClientAverageGreaterThan100"])+int(di["NggNonMSPending_ClientLowGreaterThan100"]))+'''</td>
        </tr>
        <tr>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Emergency Client</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["NggMSPending_Emergency_ClientCriticalLessThan10"]+'''</td>
            <td >'''+di["NggMSPending_Emergency_ClientHighLessThan10"]+'''</td>
            <td>'''+di["NggMSPending_Emergency_ClientAverageLessThan10"]+'''</td>
            <td>'''+di["NggMSPending_Emergency_ClientLowLessThan10"]+'''</td>
            <td>'''+str(int(di["NggMSPending_Emergency_ClientCriticalLessThan10"])+int(di["NggMSPending_Emergency_ClientHighLessThan10"])+int(di["NggMSPending_Emergency_ClientAverageLessThan10"])+int(di["NggMSPending_Emergency_ClientLowLessThan10"]))+'''</td>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Emergency Client</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["NggNonMSPending_Emergency_ClientCriticalLessThan10"]+'''</td>
            <td >'''+di["NggNonMSPending_Emergency_ClientHighLessThan10"]+'''</td>
            <td>'''+di["NggNonMSPending_Emergency_ClientAverageLessThan10"]+'''</td>
            <td>'''+di["NggNonMSPending_Emergency_ClientLowLessThan10"]+'''</td>
            <td>'''+str(int(di["NggNonMSPending_Emergency_ClientCriticalLessThan10"])+int(di["NggNonMSPending_Emergency_ClientHighLessThan10"])+int(di["NggNonMSPending_Emergency_ClientAverageLessThan10"])+int(di["NggNonMSPending_Emergency_ClientLowLessThan10"]))+'''</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["NggMSPending_Emergency_ClientCriticalGreaterThan10"]+'''</td>
            <td >'''+di["NggMSPending_Emergency_ClientHighGreaterThan10"]+'''</td>
            <td>'''+di["NggMSPending_Emergency_ClientAverageGreaterThan10"]+'''</td>
            <td>'''+di["NggMSPending_Emergency_ClientLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["NggMSPending_Emergency_ClientCriticalGreaterThan10"])+int(di["NggMSPending_Emergency_ClientHighGreaterThan10"])+int(di["NggMSPending_Emergency_ClientAverageGreaterThan10"])+int(di["NggMSPending_Emergency_ClientLowGreaterThan10"]))+'''</td>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["NggNonMSPending_Emergency_ClientCriticalGreaterThan10"]+'''</td>
            <td >'''+di["NggNonMSPending_Emergency_ClientHighGreaterThan10"]+'''</td>
            <td>'''+di["NggNonMSPending_Emergency_ClientAverageGreaterThan10"]+'''</td>
            <td>'''+di["NggNonMSPending_Emergency_ClientLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["NggNonMSPending_Emergency_ClientCriticalGreaterThan10"])+int(di["NggNonMSPending_Emergency_ClientHighGreaterThan10"])+int(di["NggNonMSPending_Emergency_ClientAverageGreaterThan10"])+int(di["NggNonMSPending_Emergency_ClientLowGreaterThan10"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["NggMSPending_Emergency_ClientCriticalGreaterThan50"]+'''</td>
            <td >'''+di["NggMSPending_Emergency_ClientHighGreaterThan50"]+'''</td>
            <td >'''+di["NggMSPending_Emergency_ClientAverageGreaterThan50"]+'''</td>
            <td >'''+di["NggMSPending_Emergency_ClientLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["NggMSPending_Emergency_ClientCriticalGreaterThan50"])+int(di["NggMSPending_Emergency_ClientHighGreaterThan50"])+int(di["NggMSPending_Emergency_ClientAverageGreaterThan50"])+int(di["NggMSPending_Emergency_ClientLowGreaterThan50"]))+'''</td>
             <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["NggNonMSPending_Emergency_ClientCriticalGreaterThan50"]+'''</td>
            <td >'''+di["NggNonMSPending_Emergency_ClientHighGreaterThan50"]+'''</td>
            <td >'''+di["NggNonMSPending_Emergency_ClientAverageGreaterThan50"]+'''</td>
            <td >'''+di["NggNonMSPending_Emergency_ClientLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["NggNonMSPending_Emergency_ClientCriticalGreaterThan50"])+int(di["NggNonMSPending_Emergency_ClientHighGreaterThan50"])+int(di["NggNonMSPending_Emergency_ClientAverageGreaterThan50"])+int(di["NggNonMSPending_Emergency_ClientLowGreaterThan50"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["NggMSPending_Emergency_ClientCriticalGreaterThan100"]+'''</td>
            <td >'''+di["NggMSPending_Emergency_ClientHighGreaterThan100"]+'''</td>
            <td >'''+di["NggMSPending_Emergency_ClientAverageGreaterThan100"]+'''</td>
            <td >'''+di["NggMSPending_Emergency_ClientLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["NggMSPending_Emergency_ClientCriticalGreaterThan100"])+int(di["NggMSPending_Emergency_ClientHighGreaterThan100"])+int(di["NggMSPending_Emergency_ClientAverageGreaterThan100"])+int(di["NggMSPending_Emergency_ClientLowGreaterThan100"]))+'''</td>
             <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["NggNonMSPending_Emergency_ClientCriticalGreaterThan100"]+'''</td>
            <td >'''+di["NggNonMSPending_Emergency_ClientHighGreaterThan100"]+'''</td>
            <td >'''+di["NggNonMSPending_Emergency_ClientAverageGreaterThan100"]+'''</td>
            <td >'''+di["NggNonMSPending_Emergency_ClientLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["NggNonMSPending_Emergency_ClientCriticalGreaterThan100"])+int(di["NggNonMSPending_Emergency_ClientHighGreaterThan100"])+int(di["NggNonMSPending_Emergency_ClientAverageGreaterThan100"])+int(di["NggNonMSPending_Emergency_ClientLowGreaterThan100"]))+'''</td>
        </tr>
        <tr>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Vendor</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["NggMSPending_VendorCriticalLessThan10"]+'''</td>
            <td >'''+di["NggMSPending_VendorHighLessThan10"]+'''</td>
            <td>'''+di["NggMSPending_VendorAverageLessThan10"]+'''</td>
            <td>'''+di["NggMSPending_VendorLowLessThan10"]+'''</td>
            <td>'''+str(int(di["NggMSPending_VendorCriticalLessThan10"])+int(di["NggMSPending_VendorHighLessThan10"])+int(di["NggMSPending_VendorAverageLessThan10"])+int(di["NggMSPending_VendorLowLessThan10"]))+'''</td>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Vendor</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["NggNonMSPending_VendorCriticalLessThan10"]+'''</td>
            <td >'''+di["NggNonMSPending_VendorHighLessThan10"]+'''</td>
            <td>'''+di["NggNonMSPending_VendorAverageLessThan10"]+'''</td>
            <td>'''+di["NggNonMSPending_VendorLowLessThan10"]+'''</td>
            <td>'''+str(int(di["NggNonMSPending_VendorCriticalLessThan10"])+int(di["NggNonMSPending_VendorHighLessThan10"])+int(di["NggNonMSPending_VendorAverageLessThan10"])+int(di["NggNonMSPending_VendorLowLessThan10"]))+'''</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["NggMSPending_VendorCriticalGreaterThan10"]+'''</td>
            <td >'''+di["NggMSPending_VendorHighGreaterThan10"]+'''</td>
            <td>'''+di["NggMSPending_VendorAverageGreaterThan10"]+'''</td>
            <td>'''+di["NggMSPending_VendorLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["NggMSPending_VendorCriticalGreaterThan10"])+int(di["NggMSPending_VendorHighGreaterThan10"])+int(di["NggMSPending_VendorAverageGreaterThan10"])+int(di["NggMSPending_VendorLowGreaterThan10"]))+'''</td>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["NggNonMSPending_VendorCriticalGreaterThan10"]+'''</td>
            <td >'''+di["NggNonMSPending_VendorHighGreaterThan10"]+'''</td>
            <td>'''+di["NggNonMSPending_VendorAverageGreaterThan10"]+'''</td>
            <td>'''+di["NggNonMSPending_VendorLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["NggNonMSPending_VendorCriticalGreaterThan10"])+int(di["NggNonMSPending_VendorHighGreaterThan10"])+int(di["NggNonMSPending_VendorAverageGreaterThan10"])+int(di["NggNonMSPending_VendorLowGreaterThan10"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["NggMSPending_VendorCriticalGreaterThan50"]+'''</td>
            <td >'''+di["NggMSPending_VendorHighGreaterThan50"]+'''</td>
            <td >'''+di["NggMSPending_VendorAverageGreaterThan50"]+'''</td>
            <td >'''+di["NggMSPending_VendorLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["NggMSPending_VendorCriticalGreaterThan50"])+int(di["NggMSPending_VendorHighGreaterThan50"])+int(di["NggMSPending_VendorAverageGreaterThan50"])+int(di["NggMSPending_VendorLowGreaterThan50"]))+'''</td>
             <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["NggNonMSPending_VendorCriticalGreaterThan50"]+'''</td>
            <td >'''+di["NggNonMSPending_VendorHighGreaterThan50"]+'''</td>
            <td >'''+di["NggNonMSPending_VendorAverageGreaterThan50"]+'''</td>
            <td >'''+di["NggNonMSPending_VendorLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["NggNonMSPending_VendorCriticalGreaterThan50"])+int(di["NggNonMSPending_VendorHighGreaterThan50"])+int(di["NggNonMSPending_VendorAverageGreaterThan50"])+int(di["NggNonMSPending_VendorLowGreaterThan50"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["NggMSPending_VendorCriticalGreaterThan100"]+'''</td>
            <td >'''+di["NggMSPending_VendorHighGreaterThan100"]+'''</td>
            <td >'''+di["NggMSPending_VendorAverageGreaterThan100"]+'''</td>
            <td >'''+di["NggMSPending_VendorLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["NggMSPending_VendorCriticalGreaterThan100"])+int(di["NggMSPending_VendorHighGreaterThan100"])+int(di["NggMSPending_VendorAverageGreaterThan100"])+int(di["NggMSPending_VendorLowGreaterThan100"]))+'''</td>
             <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["NggNonMSPending_VendorCriticalGreaterThan100"]+'''</td>
            <td >'''+di["NggNonMSPending_VendorHighGreaterThan100"]+'''</td>
            <td >'''+di["NggNonMSPending_VendorAverageGreaterThan100"]+'''</td>
            <td >'''+di["NggNonMSPending_VendorLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["NggNonMSPending_VendorCriticalGreaterThan100"])+int(di["NggNonMSPending_VendorHighGreaterThan100"])+int(di["NggNonMSPending_VendorAverageGreaterThan100"])+int(di["NggNonMSPending_VendorLowGreaterThan100"]))+'''</td>
        </tr>
        <tr>
            <td rowspan="4" bgcolor="#ADD8E6">Work In Progress</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["NggMSWork_In_ProgressCriticalLessThan10"]+'''</td>
            <td >'''+di["NggMSWork_In_ProgressHighLessThan10"]+'''</td>
            <td>'''+di["NggMSWork_In_ProgressAverageLessThan10"]+'''</td>
            <td>'''+di["NggMSWork_In_ProgressLowLessThan10"]+'''</td>
            <td>'''+str(int(di["NggMSWork_In_ProgressCriticalLessThan10"])+int(di["NggMSWork_In_ProgressHighLessThan10"])+int(di["NggMSWork_In_ProgressAverageLessThan10"])+int(di["NggMSWork_In_ProgressLowLessThan10"]))+'''</td>
            <td rowspan="4" bgcolor="#ADD8E6">Work In Progress</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["NggNonMSWork_In_ProgressCriticalLessThan10"]+'''</td>
            <td >'''+di["NggNonMSWork_In_ProgressHighLessThan10"]+'''</td>
            <td>'''+di["NggNonMSWork_In_ProgressAverageLessThan10"]+'''</td>
            <td>'''+di["NggNonMSWork_In_ProgressLowLessThan10"]+'''</td>
            <td>'''+str(int(di["NggNonMSWork_In_ProgressCriticalLessThan10"])+int(di["NggNonMSWork_In_ProgressHighLessThan10"])+int(di["NggNonMSWork_In_ProgressAverageLessThan10"])+int(di["NggNonMSWork_In_ProgressLowLessThan10"]))+'''</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["NggMSWork_In_ProgressCriticalGreaterThan10"]+'''</td>
            <td >'''+di["NggMSWork_In_ProgressHighGreaterThan10"]+'''</td>
            <td>'''+di["NggMSWork_In_ProgressAverageGreaterThan10"]+'''</td>
            <td>'''+di["NggMSWork_In_ProgressLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["NggMSWork_In_ProgressCriticalGreaterThan10"])+int(di["NggMSWork_In_ProgressHighGreaterThan10"])+int(di["NggMSWork_In_ProgressAverageGreaterThan10"])+int(di["NggMSWork_In_ProgressLowGreaterThan10"]))+'''</td>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["NggNonMSWork_In_ProgressCriticalGreaterThan10"]+'''</td>
            <td >'''+di["NggNonMSWork_In_ProgressHighGreaterThan10"]+'''</td>
            <td>'''+di["NggNonMSWork_In_ProgressAverageGreaterThan10"]+'''</td>
            <td>'''+di["NggNonMSWork_In_ProgressLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["NggNonMSWork_In_ProgressCriticalGreaterThan10"])+int(di["NggNonMSWork_In_ProgressHighGreaterThan10"])+int(di["NggNonMSWork_In_ProgressAverageGreaterThan10"])+int(di["NggNonMSWork_In_ProgressLowGreaterThan10"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["NggMSWork_In_ProgressCriticalGreaterThan50"]+'''</td>
            <td >'''+di["NggMSWork_In_ProgressHighGreaterThan50"]+'''</td>
            <td >'''+di["NggMSWork_In_ProgressAverageGreaterThan50"]+'''</td>
            <td >'''+di["NggMSWork_In_ProgressLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["NggMSWork_In_ProgressCriticalGreaterThan50"])+int(di["NggMSWork_In_ProgressHighGreaterThan50"])+int(di["NggMSWork_In_ProgressAverageGreaterThan50"])+int(di["NggMSWork_In_ProgressLowGreaterThan50"]))+'''</td>
             <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["NggNonMSWork_In_ProgressCriticalGreaterThan50"]+'''</td>
            <td >'''+di["NggNonMSWork_In_ProgressHighGreaterThan50"]+'''</td>
            <td >'''+di["NggNonMSWork_In_ProgressAverageGreaterThan50"]+'''</td>
            <td >'''+di["NggNonMSWork_In_ProgressLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["NggNonMSWork_In_ProgressCriticalGreaterThan50"])+int(di["NggNonMSWork_In_ProgressHighGreaterThan50"])+int(di["NggNonMSWork_In_ProgressAverageGreaterThan50"])+int(di["NggNonMSWork_In_ProgressLowGreaterThan50"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["NggMSWork_In_ProgressCriticalGreaterThan100"]+'''</td>
            <td >'''+di["NggMSWork_In_ProgressHighGreaterThan100"]+'''</td>
            <td >'''+di["NggMSWork_In_ProgressAverageGreaterThan100"]+'''</td>
            <td >'''+di["NggMSWork_In_ProgressLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["NggMSWork_In_ProgressCriticalGreaterThan100"])+int(di["NggMSWork_In_ProgressHighGreaterThan100"])+int(di["NggMSWork_In_ProgressAverageGreaterThan100"])+int(di["NggMSWork_In_ProgressLowGreaterThan100"]))+'''</td>
             <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["NggNonMSWork_In_ProgressCriticalGreaterThan100"]+'''</td>
            <td >'''+di["NggNonMSWork_In_ProgressHighGreaterThan100"]+'''</td>
            <td >'''+di["NggNonMSWork_In_ProgressAverageGreaterThan100"]+'''</td>
            <td >'''+di["NggNonMSWork_In_ProgressLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["NggNonMSWork_In_ProgressCriticalGreaterThan100"])+int(di["NggNonMSWork_In_ProgressHighGreaterThan100"])+int(di["NggNonMSWork_In_ProgressAverageGreaterThan100"])+int(di["NggNonMSWork_In_ProgressLowGreaterThan100"]))+'''</td>
        </tr>

        <tr class="supportgroup">
            <td colspan="14" style="text-align:center;" >US_WebMethods-EAI-lvl3</td>
        </tr>
        <tr bgcolor="#FED8B1">
            <td colspan="7" style="text-align:center;">MS Support Group</td>
            <td colspan="7" style="text-align:center;">Non MS Support Group</td>
        </tr>
        <tr>
            <td rowspan="5" bgcolor="#ADD8E6">Open</td>
            <td bgcolor="#ffcccb">Aging</td>
            <td  bgcolor="#D3D3D3">Critical</td>
            <td  bgcolor="#D3D3D3">High</td>
            <td  bgcolor="#D3D3D3">Average</td>
            <td  bgcolor="#D3D3D3">Low</td>
            <td  bgcolor="#D3D3D3">Total</td>
            <td rowspan="5" bgcolor="#ADD8E6">Open</td>
            <td bgcolor="#ffcccb">Aging</td>
            <td  bgcolor="#D3D3D3">Critical</td>
            <td  bgcolor="#D3D3D3">High</td>
            <td  bgcolor="#D3D3D3">Average</td>
            <td  bgcolor="#D3D3D3">Low</td>
            <td  bgcolor="#D3D3D3">Total</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["webMSOpenCriticalLessThan10"]+'''</td>
            <td >'''+di["webMSOpenHighLessThan10"]+'''</td>
            <td>'''+di["webMSOpenAverageLessThan10"]+'''</td>
            <td>'''+di["webMSOpenLowLessThan10"]+'''</td>
            <td>'''+str(int(di["webMSOpenCriticalLessThan10"])+int(di["webMSOpenHighLessThan10"])+int(di["webMSOpenAverageLessThan10"])+int(di["webMSOpenLowLessThan10"]))+'''</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["webNonMSOpenCriticalLessThan10"]+'''</td>
            <td >'''+di["webNonMSOpenHighLessThan10"]+'''</td>
            <td>'''+di["webNonMSOpenAverageLessThan10"]+'''</td>
            <td>'''+di["webNonMSOpenLowLessThan10"]+'''</td>
            <td>'''+str(int(di["webNonMSOpenCriticalLessThan10"])+int(di["webNonMSOpenHighLessThan10"])+int(di["webNonMSOpenAverageLessThan10"])+int(di["webNonMSOpenLowLessThan10"]))+'''</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["webMSOpenCriticalGreaterThan10"]+'''</td>
            <td >'''+di["webMSOpenHighGreaterThan10"]+'''</td>
            <td>'''+di["webMSOpenAverageGreaterThan10"]+'''</td>
            <td>'''+di["webMSOpenLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["webMSOpenCriticalGreaterThan10"])+int(di["webMSOpenHighGreaterThan10"])+int(di["webMSOpenAverageGreaterThan10"])+int(di["webMSOpenLowGreaterThan10"]))+'''</td>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["webNonMSOpenCriticalGreaterThan10"]+'''</td>
            <td >'''+di["webNonMSOpenHighGreaterThan10"]+'''</td>
            <td>'''+di["webNonMSOpenAverageGreaterThan10"]+'''</td>
            <td>'''+di["webNonMSOpenLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["webNonMSOpenCriticalGreaterThan10"])+int(di["webNonMSOpenHighGreaterThan10"])+int(di["webNonMSOpenAverageGreaterThan10"])+int(di["webNonMSOpenLowGreaterThan10"]))+'''</td>
        </tr>

        <tr>
            <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["webMSOpenCriticalGreaterThan50"]+'''</td>
            <td >'''+di["webMSOpenHighGreaterThan50"]+'''</td>
            <td >'''+di["webMSOpenAverageGreaterThan50"]+'''</td>
            <td >'''+di["webMSOpenLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["webMSOpenCriticalGreaterThan50"])+int(di["webMSOpenHighGreaterThan50"])+int(di["webMSOpenAverageGreaterThan50"])+int(di["webMSOpenLowGreaterThan50"]))+'''</td>
             <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["webNonMSOpenCriticalGreaterThan50"]+'''</td>
            <td >'''+di["webNonMSOpenHighGreaterThan50"]+'''</td>
            <td >'''+di["webNonMSOpenAverageGreaterThan50"]+'''</td>
            <td >'''+di["webNonMSOpenLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["webNonMSOpenCriticalGreaterThan50"])+int(di["webNonMSOpenHighGreaterThan50"])+int(di["webNonMSOpenAverageGreaterThan50"])+int(di["webNonMSOpenLowGreaterThan50"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["webMSOpenCriticalGreaterThan100"]+'''</td>
            <td >'''+di["webMSOpenHighGreaterThan100"]+'''</td>
            <td >'''+di["webMSOpenAverageGreaterThan100"]+'''</td>
            <td >'''+di["webMSOpenLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["webMSOpenCriticalGreaterThan100"])+int(di["webMSOpenHighGreaterThan100"])+int(di["webMSOpenAverageGreaterThan100"])+int(di["webMSOpenLowGreaterThan100"]))+'''</td>
             <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["webNonMSOpenCriticalGreaterThan100"]+'''</td>
            <td >'''+di["webNonMSOpenHighGreaterThan100"]+'''</td>
            <td >'''+di["webNonMSOpenAverageGreaterThan100"]+'''</td>
            <td >'''+di["webNonMSOpenLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["webNonMSOpenCriticalGreaterThan100"])+int(di["webNonMSOpenHighGreaterThan100"])+int(di["webNonMSOpenAverageGreaterThan100"])+int(di["webNonMSOpenLowGreaterThan100"]))+'''</td>
        </tr>

        <tr>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Client</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["webMSPending_ClientCriticalLessThan10"]+'''</td>
            <td >'''+di["webMSPending_ClientHighLessThan10"]+'''</td>
            <td>'''+di["webMSPending_ClientAverageLessThan10"]+'''</td>
            <td>'''+di["webMSPending_ClientLowLessThan10"]+'''</td>
            <td>'''+str(int(di["webMSPending_ClientCriticalLessThan10"])+int(di["webMSPending_ClientHighLessThan10"])+int(di["webMSPending_ClientAverageLessThan10"])+int(di["webMSPending_ClientLowLessThan10"]))+'''</td>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Client</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["webNonMSPending_ClientCriticalLessThan10"]+'''</td>
            <td >'''+di["webNonMSPending_ClientHighLessThan10"]+'''</td>
            <td>'''+di["webNonMSPending_ClientAverageLessThan10"]+'''</td>
            <td>'''+di["webNonMSPending_ClientLowLessThan10"]+'''</td>
            <td>'''+str(int(di["webNonMSPending_ClientCriticalLessThan10"])+int(di["webNonMSPending_ClientHighLessThan10"])+int(di["webNonMSPending_ClientAverageLessThan10"])+int(di["webNonMSPending_ClientLowLessThan10"]))+'''</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["webMSPending_ClientCriticalGreaterThan10"]+'''</td>
            <td >'''+di["webMSPending_ClientHighGreaterThan10"]+'''</td>
            <td>'''+di["webMSPending_ClientAverageGreaterThan10"]+'''</td>
            <td>'''+di["webMSPending_ClientLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["webMSPending_ClientCriticalGreaterThan10"])+int(di["webMSPending_ClientHighGreaterThan10"])+int(di["webMSPending_ClientAverageGreaterThan10"])+int(di["webMSPending_ClientLowGreaterThan10"]))+'''</td>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["webNonMSPending_ClientCriticalGreaterThan10"]+'''</td>
            <td >'''+di["webNonMSPending_ClientHighGreaterThan10"]+'''</td>
            <td>'''+di["webNonMSPending_ClientAverageGreaterThan10"]+'''</td>
            <td>'''+di["webNonMSPending_ClientLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["webNonMSPending_ClientCriticalGreaterThan10"])+int(di["webNonMSPending_ClientHighGreaterThan10"])+int(di["webNonMSPending_ClientAverageGreaterThan10"])+int(di["webNonMSPending_ClientLowGreaterThan10"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["webMSPending_ClientCriticalGreaterThan50"]+'''</td>
            <td >'''+di["webMSPending_ClientHighGreaterThan50"]+'''</td>
            <td >'''+di["webMSPending_ClientAverageGreaterThan50"]+'''</td>
            <td >'''+di["webMSPending_ClientLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["webMSPending_ClientCriticalGreaterThan50"])+int(di["webMSPending_ClientHighGreaterThan50"])+int(di["webMSPending_ClientAverageGreaterThan50"])+int(di["webMSPending_ClientLowGreaterThan50"]))+'''</td>
             <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["webNonMSPending_ClientCriticalGreaterThan50"]+'''</td>
            <td >'''+di["webNonMSPending_ClientHighGreaterThan50"]+'''</td>
            <td >'''+di["webNonMSPending_ClientAverageGreaterThan50"]+'''</td>
            <td >'''+di["webNonMSPending_ClientLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["webNonMSPending_ClientCriticalGreaterThan50"])+int(di["webNonMSPending_ClientHighGreaterThan50"])+int(di["webNonMSPending_ClientAverageGreaterThan50"])+int(di["webNonMSPending_ClientLowGreaterThan50"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["webMSPending_ClientCriticalGreaterThan100"]+'''</td>
            <td >'''+di["webMSPending_ClientHighGreaterThan100"]+'''</td>
            <td >'''+di["webMSPending_ClientAverageGreaterThan100"]+'''</td>
            <td >'''+di["webMSPending_ClientLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["webMSPending_ClientCriticalGreaterThan100"])+int(di["webMSPending_ClientHighGreaterThan100"])+int(di["webMSPending_ClientAverageGreaterThan100"])+int(di["webMSPending_ClientLowGreaterThan100"]))+'''</td>
             <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["webNonMSPending_ClientCriticalGreaterThan100"]+'''</td>
            <td >'''+di["webNonMSPending_ClientHighGreaterThan100"]+'''</td>
            <td >'''+di["webNonMSPending_ClientAverageGreaterThan100"]+'''</td>
            <td >'''+di["webNonMSPending_ClientLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["webNonMSPending_ClientCriticalGreaterThan100"])+int(di["webNonMSPending_ClientHighGreaterThan100"])+int(di["webNonMSPending_ClientAverageGreaterThan100"])+int(di["webNonMSPending_ClientLowGreaterThan100"]))+'''</td>
        </tr>
        <tr>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Emergency Client</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["webMSPending_Emergency_ClientCriticalLessThan10"]+'''</td>
            <td >'''+di["webMSPending_Emergency_ClientHighLessThan10"]+'''</td>
            <td>'''+di["webMSPending_Emergency_ClientAverageLessThan10"]+'''</td>
            <td>'''+di["webMSPending_Emergency_ClientLowLessThan10"]+'''</td>
            <td>'''+str(int(di["webMSPending_Emergency_ClientCriticalLessThan10"])+int(di["webMSPending_Emergency_ClientHighLessThan10"])+int(di["webMSPending_Emergency_ClientAverageLessThan10"])+int(di["webMSPending_Emergency_ClientLowLessThan10"]))+'''</td>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Emergency Client</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["webNonMSPending_Emergency_ClientCriticalLessThan10"]+'''</td>
            <td >'''+di["webNonMSPending_Emergency_ClientHighLessThan10"]+'''</td>
            <td>'''+di["webNonMSPending_Emergency_ClientAverageLessThan10"]+'''</td>
            <td>'''+di["webNonMSPending_Emergency_ClientLowLessThan10"]+'''</td>
            <td>'''+str(int(di["webNonMSPending_Emergency_ClientCriticalLessThan10"])+int(di["webNonMSPending_Emergency_ClientHighLessThan10"])+int(di["webNonMSPending_Emergency_ClientAverageLessThan10"])+int(di["webNonMSPending_Emergency_ClientLowLessThan10"]))+'''</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["webMSPending_Emergency_ClientCriticalGreaterThan10"]+'''</td>
            <td >'''+di["webMSPending_Emergency_ClientHighGreaterThan10"]+'''</td>
            <td>'''+di["webMSPending_Emergency_ClientAverageGreaterThan10"]+'''</td>
            <td>'''+di["webMSPending_Emergency_ClientLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["webMSPending_Emergency_ClientCriticalGreaterThan10"])+int(di["webMSPending_Emergency_ClientHighGreaterThan10"])+int(di["webMSPending_Emergency_ClientAverageGreaterThan10"])+int(di["webMSPending_Emergency_ClientLowGreaterThan10"]))+'''</td>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["webNonMSPending_Emergency_ClientCriticalGreaterThan10"]+'''</td>
            <td >'''+di["webNonMSPending_Emergency_ClientHighGreaterThan10"]+'''</td>
            <td>'''+di["webNonMSPending_Emergency_ClientAverageGreaterThan10"]+'''</td>
            <td>'''+di["webNonMSPending_Emergency_ClientLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["webNonMSPending_Emergency_ClientCriticalGreaterThan10"])+int(di["webNonMSPending_Emergency_ClientHighGreaterThan10"])+int(di["webNonMSPending_Emergency_ClientAverageGreaterThan10"])+int(di["webNonMSPending_Emergency_ClientLowGreaterThan10"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["webMSPending_Emergency_ClientCriticalGreaterThan50"]+'''</td>
            <td >'''+di["webMSPending_Emergency_ClientHighGreaterThan50"]+'''</td>
            <td >'''+di["webMSPending_Emergency_ClientAverageGreaterThan50"]+'''</td>
            <td >'''+di["webMSPending_Emergency_ClientLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["webMSPending_Emergency_ClientCriticalGreaterThan50"])+int(di["webMSPending_Emergency_ClientHighGreaterThan50"])+int(di["webMSPending_Emergency_ClientAverageGreaterThan50"])+int(di["webMSPending_Emergency_ClientLowGreaterThan50"]))+'''</td>
             <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["webNonMSPending_Emergency_ClientCriticalGreaterThan50"]+'''</td>
            <td >'''+di["webNonMSPending_Emergency_ClientHighGreaterThan50"]+'''</td>
            <td >'''+di["webNonMSPending_Emergency_ClientAverageGreaterThan50"]+'''</td>
            <td >'''+di["webNonMSPending_Emergency_ClientLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["webNonMSPending_Emergency_ClientCriticalGreaterThan50"])+int(di["webNonMSPending_Emergency_ClientHighGreaterThan50"])+int(di["webNonMSPending_Emergency_ClientAverageGreaterThan50"])+int(di["webNonMSPending_Emergency_ClientLowGreaterThan50"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["webMSPending_Emergency_ClientCriticalGreaterThan100"]+'''</td>
            <td >'''+di["webMSPending_Emergency_ClientHighGreaterThan100"]+'''</td>
            <td >'''+di["webMSPending_Emergency_ClientAverageGreaterThan100"]+'''</td>
            <td >'''+di["webMSPending_Emergency_ClientLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["webMSPending_Emergency_ClientCriticalGreaterThan100"])+int(di["webMSPending_Emergency_ClientHighGreaterThan100"])+int(di["webMSPending_Emergency_ClientAverageGreaterThan100"])+int(di["webMSPending_Emergency_ClientLowGreaterThan100"]))+'''</td>
             <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["webNonMSPending_Emergency_ClientCriticalGreaterThan100"]+'''</td>
            <td >'''+di["webNonMSPending_Emergency_ClientHighGreaterThan100"]+'''</td>
            <td >'''+di["webNonMSPending_Emergency_ClientAverageGreaterThan100"]+'''</td>
            <td >'''+di["webNonMSPending_Emergency_ClientLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["webNonMSPending_Emergency_ClientCriticalGreaterThan100"])+int(di["webNonMSPending_Emergency_ClientHighGreaterThan100"])+int(di["webNonMSPending_Emergency_ClientAverageGreaterThan100"])+int(di["webNonMSPending_Emergency_ClientLowGreaterThan100"]))+'''</td>
        </tr>
        <tr>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Vendor</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["webMSPending_VendorCriticalLessThan10"]+'''</td>
            <td >'''+di["webMSPending_VendorHighLessThan10"]+'''</td>
            <td>'''+di["webMSPending_VendorAverageLessThan10"]+'''</td>
            <td>'''+di["webMSPending_VendorLowLessThan10"]+'''</td>
            <td>'''+str(int(di["webMSPending_VendorCriticalLessThan10"])+int(di["webMSPending_VendorHighLessThan10"])+int(di["webMSPending_VendorAverageLessThan10"])+int(di["webMSPending_VendorLowLessThan10"]))+'''</td>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Vendor</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["webNonMSPending_VendorCriticalLessThan10"]+'''</td>
            <td >'''+di["webNonMSPending_VendorHighLessThan10"]+'''</td>
            <td>'''+di["webNonMSPending_VendorAverageLessThan10"]+'''</td>
            <td>'''+di["webNonMSPending_VendorLowLessThan10"]+'''</td>
            <td>'''+str(int(di["webNonMSPending_VendorCriticalLessThan10"])+int(di["webNonMSPending_VendorHighLessThan10"])+int(di["webNonMSPending_VendorAverageLessThan10"])+int(di["webNonMSPending_VendorLowLessThan10"]))+'''</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["webMSPending_VendorCriticalGreaterThan10"]+'''</td>
            <td >'''+di["webMSPending_VendorHighGreaterThan10"]+'''</td>
            <td>'''+di["webMSPending_VendorAverageGreaterThan10"]+'''</td>
            <td>'''+di["webMSPending_VendorLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["webMSPending_VendorCriticalGreaterThan10"])+int(di["webMSPending_VendorHighGreaterThan10"])+int(di["webMSPending_VendorAverageGreaterThan10"])+int(di["webMSPending_VendorLowGreaterThan10"]))+'''</td>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["webNonMSPending_VendorCriticalGreaterThan10"]+'''</td>
            <td >'''+di["webNonMSPending_VendorHighGreaterThan10"]+'''</td>
            <td>'''+di["webNonMSPending_VendorAverageGreaterThan10"]+'''</td>
            <td>'''+di["webNonMSPending_VendorLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["webNonMSPending_VendorCriticalGreaterThan10"])+int(di["webNonMSPending_VendorHighGreaterThan10"])+int(di["webNonMSPending_VendorAverageGreaterThan10"])+int(di["webNonMSPending_VendorLowGreaterThan10"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["webMSPending_VendorCriticalGreaterThan50"]+'''</td>
            <td >'''+di["webMSPending_VendorHighGreaterThan50"]+'''</td>
            <td >'''+di["webMSPending_VendorAverageGreaterThan50"]+'''</td>
            <td >'''+di["webMSPending_VendorLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["webMSPending_VendorCriticalGreaterThan50"])+int(di["webMSPending_VendorHighGreaterThan50"])+int(di["webMSPending_VendorAverageGreaterThan50"])+int(di["webMSPending_VendorLowGreaterThan50"]))+'''</td>
             <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["webNonMSPending_VendorCriticalGreaterThan50"]+'''</td>
            <td >'''+di["webNonMSPending_VendorHighGreaterThan50"]+'''</td>
            <td >'''+di["webNonMSPending_VendorAverageGreaterThan50"]+'''</td>
            <td >'''+di["webNonMSPending_VendorLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["webNonMSPending_VendorCriticalGreaterThan50"])+int(di["webNonMSPending_VendorHighGreaterThan50"])+int(di["webNonMSPending_VendorAverageGreaterThan50"])+int(di["webNonMSPending_VendorLowGreaterThan50"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["webMSPending_VendorCriticalGreaterThan100"]+'''</td>
            <td >'''+di["webMSPending_VendorHighGreaterThan100"]+'''</td>
            <td >'''+di["webMSPending_VendorAverageGreaterThan100"]+'''</td>
            <td >'''+di["webMSPending_VendorLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["webMSPending_VendorCriticalGreaterThan100"])+int(di["webMSPending_VendorHighGreaterThan100"])+int(di["webMSPending_VendorAverageGreaterThan100"])+int(di["webMSPending_VendorLowGreaterThan100"]))+'''</td>
             <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["webNonMSPending_VendorCriticalGreaterThan100"]+'''</td>
            <td >'''+di["webNonMSPending_VendorHighGreaterThan100"]+'''</td>
            <td >'''+di["webNonMSPending_VendorAverageGreaterThan100"]+'''</td>
            <td >'''+di["webNonMSPending_VendorLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["webNonMSPending_VendorCriticalGreaterThan100"])+int(di["webNonMSPending_VendorHighGreaterThan100"])+int(di["webNonMSPending_VendorAverageGreaterThan100"])+int(di["webNonMSPending_VendorLowGreaterThan100"]))+'''</td>
        </tr>
        <tr>
            <td rowspan="4" bgcolor="#ADD8E6">Work In Progress</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["webMSWork_In_ProgressCriticalLessThan10"]+'''</td>
            <td >'''+di["webMSWork_In_ProgressHighLessThan10"]+'''</td>
            <td>'''+di["webMSWork_In_ProgressAverageLessThan10"]+'''</td>
            <td>'''+di["webMSWork_In_ProgressLowLessThan10"]+'''</td>
            <td>'''+str(int(di["webMSWork_In_ProgressCriticalLessThan10"])+int(di["webMSWork_In_ProgressHighLessThan10"])+int(di["webMSWork_In_ProgressAverageLessThan10"])+int(di["webMSWork_In_ProgressLowLessThan10"]))+'''</td>
            <td rowspan="4" bgcolor="#ADD8E6">Work In Progress</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["webNonMSWork_In_ProgressCriticalLessThan10"]+'''</td>
            <td >'''+di["webNonMSWork_In_ProgressHighLessThan10"]+'''</td>
            <td>'''+di["webNonMSWork_In_ProgressAverageLessThan10"]+'''</td>
            <td>'''+di["webNonMSWork_In_ProgressLowLessThan10"]+'''</td>
            <td>'''+str(int(di["webNonMSWork_In_ProgressCriticalLessThan10"])+int(di["webNonMSWork_In_ProgressHighLessThan10"])+int(di["webNonMSWork_In_ProgressAverageLessThan10"])+int(di["webNonMSWork_In_ProgressLowLessThan10"]))+'''</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["webMSWork_In_ProgressCriticalGreaterThan10"]+'''</td>
            <td >'''+di["webMSWork_In_ProgressHighGreaterThan10"]+'''</td>
            <td>'''+di["webMSWork_In_ProgressAverageGreaterThan10"]+'''</td>
            <td>'''+di["webMSWork_In_ProgressLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["webMSWork_In_ProgressCriticalGreaterThan10"])+int(di["webMSWork_In_ProgressHighGreaterThan10"])+int(di["webMSWork_In_ProgressAverageGreaterThan10"])+int(di["webMSWork_In_ProgressLowGreaterThan10"]))+'''</td>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["webNonMSWork_In_ProgressCriticalGreaterThan10"]+'''</td>
            <td >'''+di["webNonMSWork_In_ProgressHighGreaterThan10"]+'''</td>
            <td>'''+di["webNonMSWork_In_ProgressAverageGreaterThan10"]+'''</td>
            <td>'''+di["webNonMSWork_In_ProgressLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["webNonMSWork_In_ProgressCriticalGreaterThan10"])+int(di["webNonMSWork_In_ProgressHighGreaterThan10"])+int(di["webNonMSWork_In_ProgressAverageGreaterThan10"])+int(di["webNonMSWork_In_ProgressLowGreaterThan10"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["webMSWork_In_ProgressCriticalGreaterThan50"]+'''</td>
            <td >'''+di["webMSWork_In_ProgressHighGreaterThan50"]+'''</td>
            <td >'''+di["webMSWork_In_ProgressAverageGreaterThan50"]+'''</td>
            <td >'''+di["webMSWork_In_ProgressLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["webMSWork_In_ProgressCriticalGreaterThan50"])+int(di["webMSWork_In_ProgressHighGreaterThan50"])+int(di["webMSWork_In_ProgressAverageGreaterThan50"])+int(di["webMSWork_In_ProgressLowGreaterThan50"]))+'''</td>
             <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["webNonMSWork_In_ProgressCriticalGreaterThan50"]+'''</td>
            <td >'''+di["webNonMSWork_In_ProgressHighGreaterThan50"]+'''</td>
            <td >'''+di["webNonMSWork_In_ProgressAverageGreaterThan50"]+'''</td>
            <td >'''+di["webNonMSWork_In_ProgressLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["webNonMSWork_In_ProgressCriticalGreaterThan50"])+int(di["webNonMSWork_In_ProgressHighGreaterThan50"])+int(di["webNonMSWork_In_ProgressAverageGreaterThan50"])+int(di["webNonMSWork_In_ProgressLowGreaterThan50"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["webMSWork_In_ProgressCriticalGreaterThan100"]+'''</td>
            <td >'''+di["webMSWork_In_ProgressHighGreaterThan100"]+'''</td>
            <td >'''+di["webMSWork_In_ProgressAverageGreaterThan100"]+'''</td>
            <td >'''+di["webMSWork_In_ProgressLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["webMSWork_In_ProgressCriticalGreaterThan100"])+int(di["webMSWork_In_ProgressHighGreaterThan100"])+int(di["webMSWork_In_ProgressAverageGreaterThan100"])+int(di["webMSWork_In_ProgressLowGreaterThan100"]))+'''</td>
             <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["webNonMSWork_In_ProgressCriticalGreaterThan100"]+'''</td>
            <td >'''+di["webNonMSWork_In_ProgressHighGreaterThan100"]+'''</td>
            <td >'''+di["webNonMSWork_In_ProgressAverageGreaterThan100"]+'''</td>
            <td >'''+di["webNonMSWork_In_ProgressLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["webNonMSWork_In_ProgressCriticalGreaterThan100"])+int(di["webNonMSWork_In_ProgressHighGreaterThan100"])+int(di["webNonMSWork_In_ProgressAverageGreaterThan100"])+int(di["webNonMSWork_In_ProgressLowGreaterThan100"]))+'''</td>
        </tr>
        <tr class="supportgroup" >
            <td colspan="14" style="text-align:center;" >US_EDI-Analysts</td>
        </tr>
        <tr bgcolor="#FED8B1">
            <td colspan="7" style="text-align:center;">MS Support Group</td>
            <td colspan="7" style="text-align:center;">Non MS Support Group</td>
        </tr>
        <tr>
            <td rowspan="5" bgcolor="#ADD8E6">Open</td>
            <td bgcolor="#ffcccb">Aging</td>
            <td  bgcolor="#D3D3D3">Critical</td>
            <td  bgcolor="#D3D3D3">High</td>
            <td  bgcolor="#D3D3D3">Average</td>
            <td  bgcolor="#D3D3D3">Low</td>
            <td  bgcolor="#D3D3D3">Total</td>
            <td rowspan="5" bgcolor="#ADD8E6">Open</td>
            <td bgcolor="#ffcccb">Aging</td>
            <td  bgcolor="#D3D3D3">Critical</td>
            <td  bgcolor="#D3D3D3">High</td>
            <td  bgcolor="#D3D3D3">Average</td>
            <td  bgcolor="#D3D3D3">Low</td>
            <td  bgcolor="#D3D3D3">Total</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["eanMSOpenCriticalLessThan10"]+'''</td>
            <td >'''+di["eanMSOpenHighLessThan10"]+'''</td>
            <td>'''+di["eanMSOpenAverageLessThan10"]+'''</td>
            <td>'''+di["eanMSOpenLowLessThan10"]+'''</td>
            <td>'''+str(int(di["eanMSOpenCriticalLessThan10"])+int(di["eanMSOpenHighLessThan10"])+int(di["eanMSOpenAverageLessThan10"])+int(di["eanMSOpenLowLessThan10"]))+'''</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["eanNonMSOpenCriticalLessThan10"]+'''</td>
            <td >'''+di["eanNonMSOpenHighLessThan10"]+'''</td>
            <td>'''+di["eanNonMSOpenAverageLessThan10"]+'''</td>
            <td>'''+di["eanNonMSOpenLowLessThan10"]+'''</td>
            <td>'''+str(int(di["eanNonMSOpenCriticalLessThan10"])+int(di["eanNonMSOpenHighLessThan10"])+int(di["eanNonMSOpenAverageLessThan10"])+int(di["eanNonMSOpenLowLessThan10"]))+'''</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["eanMSOpenCriticalGreaterThan10"]+'''</td>
            <td >'''+di["eanMSOpenHighGreaterThan10"]+'''</td>
            <td>'''+di["eanMSOpenAverageGreaterThan10"]+'''</td>
            <td>'''+di["eanMSOpenLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["eanMSOpenCriticalGreaterThan10"])+int(di["eanMSOpenHighGreaterThan10"])+int(di["eanMSOpenAverageGreaterThan10"])+int(di["eanMSOpenLowGreaterThan10"]))+'''</td>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["eanNonMSOpenCriticalGreaterThan10"]+'''</td>
            <td >'''+di["eanNonMSOpenHighGreaterThan10"]+'''</td>
            <td>'''+di["eanNonMSOpenAverageGreaterThan10"]+'''</td>
            <td>'''+di["eanNonMSOpenLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["eanNonMSOpenCriticalGreaterThan10"])+int(di["eanNonMSOpenHighGreaterThan10"])+int(di["eanNonMSOpenAverageGreaterThan10"])+int(di["eanNonMSOpenLowGreaterThan10"]))+'''</td>
        </tr>

        <tr>
            <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["eanMSOpenCriticalGreaterThan50"]+'''</td>
            <td >'''+di["eanMSOpenHighGreaterThan50"]+'''</td>
            <td >'''+di["eanMSOpenAverageGreaterThan50"]+'''</td>
            <td >'''+di["eanMSOpenLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["eanMSOpenCriticalGreaterThan50"])+int(di["eanMSOpenHighGreaterThan50"])+int(di["eanMSOpenAverageGreaterThan50"])+int(di["eanMSOpenLowGreaterThan50"]))+'''</td>
             <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["eanNonMSOpenCriticalGreaterThan50"]+'''</td>
            <td >'''+di["eanNonMSOpenHighGreaterThan50"]+'''</td>
            <td >'''+di["eanNonMSOpenAverageGreaterThan50"]+'''</td>
            <td >'''+di["eanNonMSOpenLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["eanNonMSOpenCriticalGreaterThan50"])+int(di["eanNonMSOpenHighGreaterThan50"])+int(di["eanNonMSOpenAverageGreaterThan50"])+int(di["eanNonMSOpenLowGreaterThan50"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["eanMSOpenCriticalGreaterThan100"]+'''</td>
            <td >'''+di["eanMSOpenHighGreaterThan100"]+'''</td>
            <td >'''+di["eanMSOpenAverageGreaterThan100"]+'''</td>
            <td >'''+di["eanMSOpenLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["eanMSOpenCriticalGreaterThan100"])+int(di["eanMSOpenHighGreaterThan100"])+int(di["eanMSOpenAverageGreaterThan100"])+int(di["eanMSOpenLowGreaterThan100"]))+'''</td>
             <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["eanNonMSOpenCriticalGreaterThan100"]+'''</td>
            <td >'''+di["eanNonMSOpenHighGreaterThan100"]+'''</td>
            <td >'''+di["eanNonMSOpenAverageGreaterThan100"]+'''</td>
            <td >'''+di["eanNonMSOpenLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["eanNonMSOpenCriticalGreaterThan100"])+int(di["eanNonMSOpenHighGreaterThan100"])+int(di["eanNonMSOpenAverageGreaterThan100"])+int(di["eanNonMSOpenLowGreaterThan100"]))+'''</td>
        </tr>

        <tr>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Client</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["eanMSPending_ClientCriticalLessThan10"]+'''</td>
            <td >'''+di["eanMSPending_ClientHighLessThan10"]+'''</td>
            <td>'''+di["eanMSPending_ClientAverageLessThan10"]+'''</td>
            <td>'''+di["eanMSPending_ClientLowLessThan10"]+'''</td>
            <td>'''+str(int(di["eanMSPending_ClientCriticalLessThan10"])+int(di["eanMSPending_ClientHighLessThan10"])+int(di["eanMSPending_ClientAverageLessThan10"])+int(di["eanMSPending_ClientLowLessThan10"]))+'''</td>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Client</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["eanNonMSPending_ClientCriticalLessThan10"]+'''</td>
            <td >'''+di["eanNonMSPending_ClientHighLessThan10"]+'''</td>
            <td>'''+di["eanNonMSPending_ClientAverageLessThan10"]+'''</td>
            <td>'''+di["eanNonMSPending_ClientLowLessThan10"]+'''</td>
            <td>'''+str(int(di["eanNonMSPending_ClientCriticalLessThan10"])+int(di["eanNonMSPending_ClientHighLessThan10"])+int(di["eanNonMSPending_ClientAverageLessThan10"])+int(di["eanNonMSPending_ClientLowLessThan10"]))+'''</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["eanMSPending_ClientCriticalGreaterThan10"]+'''</td>
            <td >'''+di["eanMSPending_ClientHighGreaterThan10"]+'''</td>
            <td>'''+di["eanMSPending_ClientAverageGreaterThan10"]+'''</td>
            <td>'''+di["eanMSPending_ClientLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["eanMSPending_ClientCriticalGreaterThan10"])+int(di["eanMSPending_ClientHighGreaterThan10"])+int(di["eanMSPending_ClientAverageGreaterThan10"])+int(di["eanMSPending_ClientLowGreaterThan10"]))+'''</td>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["eanNonMSPending_ClientCriticalGreaterThan10"]+'''</td>
            <td >'''+di["eanNonMSPending_ClientHighGreaterThan10"]+'''</td>
            <td>'''+di["eanNonMSPending_ClientAverageGreaterThan10"]+'''</td>
            <td>'''+di["eanNonMSPending_ClientLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["eanNonMSPending_ClientCriticalGreaterThan10"])+int(di["eanNonMSPending_ClientHighGreaterThan10"])+int(di["eanNonMSPending_ClientAverageGreaterThan10"])+int(di["eanNonMSPending_ClientLowGreaterThan10"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["eanMSPending_ClientCriticalGreaterThan50"]+'''</td>
            <td >'''+di["eanMSPending_ClientHighGreaterThan50"]+'''</td>
            <td >'''+di["eanMSPending_ClientAverageGreaterThan50"]+'''</td>
            <td >'''+di["eanMSPending_ClientLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["eanMSPending_ClientCriticalGreaterThan50"])+int(di["eanMSPending_ClientHighGreaterThan50"])+int(di["eanMSPending_ClientAverageGreaterThan50"])+int(di["eanMSPending_ClientLowGreaterThan50"]))+'''</td>
             <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["eanNonMSPending_ClientCriticalGreaterThan50"]+'''</td>
            <td >'''+di["eanNonMSPending_ClientHighGreaterThan50"]+'''</td>
            <td >'''+di["eanNonMSPending_ClientAverageGreaterThan50"]+'''</td>
            <td >'''+di["eanNonMSPending_ClientLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["eanNonMSPending_ClientCriticalGreaterThan50"])+int(di["eanNonMSPending_ClientHighGreaterThan50"])+int(di["eanNonMSPending_ClientAverageGreaterThan50"])+int(di["eanNonMSPending_ClientLowGreaterThan50"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["eanMSPending_ClientCriticalGreaterThan100"]+'''</td>
            <td >'''+di["eanMSPending_ClientHighGreaterThan100"]+'''</td>
            <td >'''+di["eanMSPending_ClientAverageGreaterThan100"]+'''</td>
            <td >'''+di["eanMSPending_ClientLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["eanMSPending_ClientCriticalGreaterThan100"])+int(di["eanMSPending_ClientHighGreaterThan100"])+int(di["eanMSPending_ClientAverageGreaterThan100"])+int(di["eanMSPending_ClientLowGreaterThan100"]))+'''</td>
             <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["eanNonMSPending_ClientCriticalGreaterThan100"]+'''</td>
            <td >'''+di["eanNonMSPending_ClientHighGreaterThan100"]+'''</td>
            <td >'''+di["eanNonMSPending_ClientAverageGreaterThan100"]+'''</td>
            <td >'''+di["eanNonMSPending_ClientLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["eanNonMSPending_ClientCriticalGreaterThan100"])+int(di["eanNonMSPending_ClientHighGreaterThan100"])+int(di["eanNonMSPending_ClientAverageGreaterThan100"])+int(di["eanNonMSPending_ClientLowGreaterThan100"]))+'''</td>
        </tr>
        <tr>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Emergency Client</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["eanMSPending_Emergency_ClientCriticalLessThan10"]+'''</td>
            <td >'''+di["eanMSPending_Emergency_ClientHighLessThan10"]+'''</td>
            <td>'''+di["eanMSPending_Emergency_ClientAverageLessThan10"]+'''</td>
            <td>'''+di["eanMSPending_Emergency_ClientLowLessThan10"]+'''</td>
            <td>'''+str(int(di["eanMSPending_Emergency_ClientCriticalLessThan10"])+int(di["eanMSPending_Emergency_ClientHighLessThan10"])+int(di["eanMSPending_Emergency_ClientAverageLessThan10"])+int(di["eanMSPending_Emergency_ClientLowLessThan10"]))+'''</td>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Emergency Client</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["eanNonMSPending_Emergency_ClientCriticalLessThan10"]+'''</td>
            <td >'''+di["eanNonMSPending_Emergency_ClientHighLessThan10"]+'''</td>
            <td>'''+di["eanNonMSPending_Emergency_ClientAverageLessThan10"]+'''</td>
            <td>'''+di["eanNonMSPending_Emergency_ClientLowLessThan10"]+'''</td>
            <td>'''+str(int(di["eanNonMSPending_Emergency_ClientCriticalLessThan10"])+int(di["eanNonMSPending_Emergency_ClientHighLessThan10"])+int(di["eanNonMSPending_Emergency_ClientAverageLessThan10"])+int(di["eanNonMSPending_Emergency_ClientLowLessThan10"]))+'''</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["eanMSPending_Emergency_ClientCriticalGreaterThan10"]+'''</td>
            <td >'''+di["eanMSPending_Emergency_ClientHighGreaterThan10"]+'''</td>
            <td>'''+di["eanMSPending_Emergency_ClientAverageGreaterThan10"]+'''</td>
            <td>'''+di["eanMSPending_Emergency_ClientLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["eanMSPending_Emergency_ClientCriticalGreaterThan10"])+int(di["eanMSPending_Emergency_ClientHighGreaterThan10"])+int(di["eanMSPending_Emergency_ClientAverageGreaterThan10"])+int(di["eanMSPending_Emergency_ClientLowGreaterThan10"]))+'''</td>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["eanNonMSPending_Emergency_ClientCriticalGreaterThan10"]+'''</td>
            <td >'''+di["eanNonMSPending_Emergency_ClientHighGreaterThan10"]+'''</td>
            <td>'''+di["eanNonMSPending_Emergency_ClientAverageGreaterThan10"]+'''</td>
            <td>'''+di["eanNonMSPending_Emergency_ClientLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["eanNonMSPending_Emergency_ClientCriticalGreaterThan10"])+int(di["eanNonMSPending_Emergency_ClientHighGreaterThan10"])+int(di["eanNonMSPending_Emergency_ClientAverageGreaterThan10"])+int(di["eanNonMSPending_Emergency_ClientLowGreaterThan10"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["eanMSPending_Emergency_ClientCriticalGreaterThan50"]+'''</td>
            <td >'''+di["eanMSPending_Emergency_ClientHighGreaterThan50"]+'''</td>
            <td >'''+di["eanMSPending_Emergency_ClientAverageGreaterThan50"]+'''</td>
            <td >'''+di["eanMSPending_Emergency_ClientLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["eanMSPending_Emergency_ClientCriticalGreaterThan50"])+int(di["eanMSPending_Emergency_ClientHighGreaterThan50"])+int(di["eanMSPending_Emergency_ClientAverageGreaterThan50"])+int(di["eanMSPending_Emergency_ClientLowGreaterThan50"]))+'''</td>
             <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["eanNonMSPending_Emergency_ClientCriticalGreaterThan50"]+'''</td>
            <td >'''+di["eanNonMSPending_Emergency_ClientHighGreaterThan50"]+'''</td>
            <td >'''+di["eanNonMSPending_Emergency_ClientAverageGreaterThan50"]+'''</td>
            <td >'''+di["eanNonMSPending_Emergency_ClientLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["eanNonMSPending_Emergency_ClientCriticalGreaterThan50"])+int(di["eanNonMSPending_Emergency_ClientHighGreaterThan50"])+int(di["eanNonMSPending_Emergency_ClientAverageGreaterThan50"])+int(di["eanNonMSPending_Emergency_ClientLowGreaterThan50"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["eanMSPending_Emergency_ClientCriticalGreaterThan100"]+'''</td>
            <td >'''+di["eanMSPending_Emergency_ClientHighGreaterThan100"]+'''</td>
            <td >'''+di["eanMSPending_Emergency_ClientAverageGreaterThan100"]+'''</td>
            <td >'''+di["eanMSPending_Emergency_ClientLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["eanMSPending_Emergency_ClientCriticalGreaterThan100"])+int(di["eanMSPending_Emergency_ClientHighGreaterThan100"])+int(di["eanMSPending_Emergency_ClientAverageGreaterThan100"])+int(di["eanMSPending_Emergency_ClientLowGreaterThan100"]))+'''</td>
             <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["eanNonMSPending_Emergency_ClientCriticalGreaterThan100"]+'''</td>
            <td >'''+di["eanNonMSPending_Emergency_ClientHighGreaterThan100"]+'''</td>
            <td >'''+di["eanNonMSPending_Emergency_ClientAverageGreaterThan100"]+'''</td>
            <td >'''+di["eanNonMSPending_Emergency_ClientLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["eanNonMSPending_Emergency_ClientCriticalGreaterThan100"])+int(di["eanNonMSPending_Emergency_ClientHighGreaterThan100"])+int(di["eanNonMSPending_Emergency_ClientAverageGreaterThan100"])+int(di["eanNonMSPending_Emergency_ClientLowGreaterThan100"]))+'''</td>
        </tr>
        <tr>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Vendor</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["eanMSPending_VendorCriticalLessThan10"]+'''</td>
            <td >'''+di["eanMSPending_VendorHighLessThan10"]+'''</td>
            <td>'''+di["eanMSPending_VendorAverageLessThan10"]+'''</td>
            <td>'''+di["eanMSPending_VendorLowLessThan10"]+'''</td>
            <td>'''+str(int(di["eanMSPending_VendorCriticalLessThan10"])+int(di["eanMSPending_VendorHighLessThan10"])+int(di["eanMSPending_VendorAverageLessThan10"])+int(di["eanMSPending_VendorLowLessThan10"]))+'''</td>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Vendor</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["eanNonMSPending_VendorCriticalLessThan10"]+'''</td>
            <td >'''+di["eanNonMSPending_VendorHighLessThan10"]+'''</td>
            <td>'''+di["eanNonMSPending_VendorAverageLessThan10"]+'''</td>
            <td>'''+di["eanNonMSPending_VendorLowLessThan10"]+'''</td>
            <td>'''+str(int(di["eanNonMSPending_VendorCriticalLessThan10"])+int(di["eanNonMSPending_VendorHighLessThan10"])+int(di["eanNonMSPending_VendorAverageLessThan10"])+int(di["eanNonMSPending_VendorLowLessThan10"]))+'''</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["eanMSPending_VendorCriticalGreaterThan10"]+'''</td>
            <td >'''+di["eanMSPending_VendorHighGreaterThan10"]+'''</td>
            <td>'''+di["eanMSPending_VendorAverageGreaterThan10"]+'''</td>
            <td>'''+di["eanMSPending_VendorLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["eanMSPending_VendorCriticalGreaterThan10"])+int(di["eanMSPending_VendorHighGreaterThan10"])+int(di["eanMSPending_VendorAverageGreaterThan10"])+int(di["eanMSPending_VendorLowGreaterThan10"]))+'''</td>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["eanNonMSPending_VendorCriticalGreaterThan10"]+'''</td>
            <td >'''+di["eanNonMSPending_VendorHighGreaterThan10"]+'''</td>
            <td>'''+di["eanNonMSPending_VendorAverageGreaterThan10"]+'''</td>
            <td>'''+di["eanNonMSPending_VendorLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["eanNonMSPending_VendorCriticalGreaterThan10"])+int(di["eanNonMSPending_VendorHighGreaterThan10"])+int(di["eanNonMSPending_VendorAverageGreaterThan10"])+int(di["eanNonMSPending_VendorLowGreaterThan10"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["eanMSPending_VendorCriticalGreaterThan50"]+'''</td>
            <td >'''+di["eanMSPending_VendorHighGreaterThan50"]+'''</td>
            <td >'''+di["eanMSPending_VendorAverageGreaterThan50"]+'''</td>
            <td >'''+di["eanMSPending_VendorLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["eanMSPending_VendorCriticalGreaterThan50"])+int(di["eanMSPending_VendorHighGreaterThan50"])+int(di["eanMSPending_VendorAverageGreaterThan50"])+int(di["eanMSPending_VendorLowGreaterThan50"]))+'''</td>
             <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["eanNonMSPending_VendorCriticalGreaterThan50"]+'''</td>
            <td >'''+di["eanNonMSPending_VendorHighGreaterThan50"]+'''</td>
            <td >'''+di["eanNonMSPending_VendorAverageGreaterThan50"]+'''</td>
            <td >'''+di["eanNonMSPending_VendorLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["eanNonMSPending_VendorCriticalGreaterThan50"])+int(di["eanNonMSPending_VendorHighGreaterThan50"])+int(di["eanNonMSPending_VendorAverageGreaterThan50"])+int(di["eanNonMSPending_VendorLowGreaterThan50"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["eanMSPending_VendorCriticalGreaterThan100"]+'''</td>
            <td >'''+di["eanMSPending_VendorHighGreaterThan100"]+'''</td>
            <td >'''+di["eanMSPending_VendorAverageGreaterThan100"]+'''</td>
            <td >'''+di["eanMSPending_VendorLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["eanMSPending_VendorCriticalGreaterThan100"])+int(di["eanMSPending_VendorHighGreaterThan100"])+int(di["eanMSPending_VendorAverageGreaterThan100"])+int(di["eanMSPending_VendorLowGreaterThan100"]))+'''</td>
             <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["eanNonMSPending_VendorCriticalGreaterThan100"]+'''</td>
            <td >'''+di["eanNonMSPending_VendorHighGreaterThan100"]+'''</td>
            <td >'''+di["eanNonMSPending_VendorAverageGreaterThan100"]+'''</td>
            <td >'''+di["eanNonMSPending_VendorLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["eanNonMSPending_VendorCriticalGreaterThan100"])+int(di["eanNonMSPending_VendorHighGreaterThan100"])+int(di["eanNonMSPending_VendorAverageGreaterThan100"])+int(di["eanNonMSPending_VendorLowGreaterThan100"]))+'''</td>
        </tr>
        <tr>
            <td rowspan="4" bgcolor="#ADD8E6">Work In Progress</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["eanMSWork_In_ProgressCriticalLessThan10"]+'''</td>
            <td >'''+di["eanMSWork_In_ProgressHighLessThan10"]+'''</td>
            <td>'''+di["eanMSWork_In_ProgressAverageLessThan10"]+'''</td>
            <td>'''+di["eanMSWork_In_ProgressLowLessThan10"]+'''</td>
            <td>'''+str(int(di["eanMSWork_In_ProgressCriticalLessThan10"])+int(di["eanMSWork_In_ProgressHighLessThan10"])+int(di["eanMSWork_In_ProgressAverageLessThan10"])+int(di["eanMSWork_In_ProgressLowLessThan10"]))+'''</td>
            <td rowspan="4" bgcolor="#ADD8E6">Work In Progress</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["eanNonMSWork_In_ProgressCriticalLessThan10"]+'''</td>
            <td >'''+di["eanNonMSWork_In_ProgressHighLessThan10"]+'''</td>
            <td>'''+di["eanNonMSWork_In_ProgressAverageLessThan10"]+'''</td>
            <td>'''+di["eanNonMSWork_In_ProgressLowLessThan10"]+'''</td>
            <td>'''+str(int(di["eanNonMSWork_In_ProgressCriticalLessThan10"])+int(di["eanNonMSWork_In_ProgressHighLessThan10"])+int(di["eanNonMSWork_In_ProgressAverageLessThan10"])+int(di["eanNonMSWork_In_ProgressLowLessThan10"]))+'''</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["eanMSWork_In_ProgressCriticalGreaterThan10"]+'''</td>
            <td >'''+di["eanMSWork_In_ProgressHighGreaterThan10"]+'''</td>
            <td>'''+di["eanMSWork_In_ProgressAverageGreaterThan10"]+'''</td>
            <td>'''+di["eanMSWork_In_ProgressLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["eanMSWork_In_ProgressCriticalGreaterThan10"])+int(di["eanMSWork_In_ProgressHighGreaterThan10"])+int(di["eanMSWork_In_ProgressAverageGreaterThan10"])+int(di["eanMSWork_In_ProgressLowGreaterThan10"]))+'''</td>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["eanNonMSWork_In_ProgressCriticalGreaterThan10"]+'''</td>
            <td >'''+di["eanNonMSWork_In_ProgressHighGreaterThan10"]+'''</td>
            <td>'''+di["eanNonMSWork_In_ProgressAverageGreaterThan10"]+'''</td>
            <td>'''+di["eanNonMSWork_In_ProgressLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["eanNonMSWork_In_ProgressCriticalGreaterThan10"])+int(di["eanNonMSWork_In_ProgressHighGreaterThan10"])+int(di["eanNonMSWork_In_ProgressAverageGreaterThan10"])+int(di["eanNonMSWork_In_ProgressLowGreaterThan10"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["eanMSWork_In_ProgressCriticalGreaterThan50"]+'''</td>
            <td >'''+di["eanMSWork_In_ProgressHighGreaterThan50"]+'''</td>
            <td >'''+di["eanMSWork_In_ProgressAverageGreaterThan50"]+'''</td>
            <td >'''+di["eanMSWork_In_ProgressLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["eanMSWork_In_ProgressCriticalGreaterThan50"])+int(di["eanMSWork_In_ProgressHighGreaterThan50"])+int(di["eanMSWork_In_ProgressAverageGreaterThan50"])+int(di["eanMSWork_In_ProgressLowGreaterThan50"]))+'''</td>
             <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["eanNonMSWork_In_ProgressCriticalGreaterThan50"]+'''</td>
            <td >'''+di["eanNonMSWork_In_ProgressHighGreaterThan50"]+'''</td>
            <td >'''+di["eanNonMSWork_In_ProgressAverageGreaterThan50"]+'''</td>
            <td >'''+di["eanNonMSWork_In_ProgressLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["eanNonMSWork_In_ProgressCriticalGreaterThan50"])+int(di["eanNonMSWork_In_ProgressHighGreaterThan50"])+int(di["eanNonMSWork_In_ProgressAverageGreaterThan50"])+int(di["eanNonMSWork_In_ProgressLowGreaterThan50"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["eanMSWork_In_ProgressCriticalGreaterThan100"]+'''</td>
            <td >'''+di["eanMSWork_In_ProgressHighGreaterThan100"]+'''</td>
            <td >'''+di["eanMSWork_In_ProgressAverageGreaterThan100"]+'''</td>
            <td >'''+di["eanMSWork_In_ProgressLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["eanMSWork_In_ProgressCriticalGreaterThan100"])+int(di["eanMSWork_In_ProgressHighGreaterThan100"])+int(di["eanMSWork_In_ProgressAverageGreaterThan100"])+int(di["eanMSWork_In_ProgressLowGreaterThan100"]))+'''</td>
             <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["eanNonMSWork_In_ProgressCriticalGreaterThan100"]+'''</td>
            <td >'''+di["eanNonMSWork_In_ProgressHighGreaterThan100"]+'''</td>
            <td >'''+di["eanNonMSWork_In_ProgressAverageGreaterThan100"]+'''</td>
            <td >'''+di["eanNonMSWork_In_ProgressLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["eanNonMSWork_In_ProgressCriticalGreaterThan100"])+int(di["eanNonMSWork_In_ProgressHighGreaterThan100"])+int(di["eanNonMSWork_In_ProgressAverageGreaterThan100"])+int(di["eanNonMSWork_In_ProgressLowGreaterThan100"]))+'''</td>
        </tr>

        <tr class="supportgroup" >
            <td colspan="14" style="text-align:center;" >US_EDI-Support</td>
        </tr>
        <tr bgcolor="#FED8B1">
            <td colspan="7" style="text-align:center;">MS Support Group</td>
            <td colspan="7" style="text-align:center;">Non MS Support Group</td>
        </tr>
        <tr>
            <td rowspan="5" bgcolor="#ADD8E6">Open</td>
            <td bgcolor="#ffcccb">Aging</td>
            <td  bgcolor="#D3D3D3">Critical</td>
            <td  bgcolor="#D3D3D3">High</td>
            <td  bgcolor="#D3D3D3">Average</td>
            <td  bgcolor="#D3D3D3">Low</td>
            <td  bgcolor="#D3D3D3">Total</td>
            <td rowspan="5" bgcolor="#ADD8E6">Open</td>
            <td bgcolor="#ffcccb">Aging</td>
            <td  bgcolor="#D3D3D3">Critical</td>
            <td  bgcolor="#D3D3D3">High</td>
            <td  bgcolor="#D3D3D3">Average</td>
            <td  bgcolor="#D3D3D3">Low</td>
            <td  bgcolor="#D3D3D3">Total</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["edsuMSOpenCriticalLessThan10"]+'''</td>
            <td >'''+di["edsuMSOpenHighLessThan10"]+'''</td>
            <td>'''+di["edsuMSOpenAverageLessThan10"]+'''</td>
            <td>'''+di["edsuMSOpenLowLessThan10"]+'''</td>
            <td>'''+str(int(di["edsuMSOpenCriticalLessThan10"])+int(di["edsuMSOpenHighLessThan10"])+int(di["edsuMSOpenAverageLessThan10"])+int(di["edsuMSOpenLowLessThan10"]))+'''</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["edsuNonMSOpenCriticalLessThan10"]+'''</td>
            <td >'''+di["edsuNonMSOpenHighLessThan10"]+'''</td>
            <td>'''+di["edsuNonMSOpenAverageLessThan10"]+'''</td>
            <td>'''+di["edsuNonMSOpenLowLessThan10"]+'''</td>
            <td>'''+str(int(di["edsuNonMSOpenCriticalLessThan10"])+int(di["edsuNonMSOpenHighLessThan10"])+int(di["edsuNonMSOpenAverageLessThan10"])+int(di["edsuNonMSOpenLowLessThan10"]))+'''</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["edsuMSOpenCriticalGreaterThan10"]+'''</td>
            <td >'''+di["edsuMSOpenHighGreaterThan10"]+'''</td>
            <td>'''+di["edsuMSOpenAverageGreaterThan10"]+'''</td>
            <td>'''+di["edsuMSOpenLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["edsuMSOpenCriticalGreaterThan10"])+int(di["edsuMSOpenHighGreaterThan10"])+int(di["edsuMSOpenAverageGreaterThan10"])+int(di["edsuMSOpenLowGreaterThan10"]))+'''</td>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["edsuNonMSOpenCriticalGreaterThan10"]+'''</td>
            <td >'''+di["edsuNonMSOpenHighGreaterThan10"]+'''</td>
            <td>'''+di["edsuNonMSOpenAverageGreaterThan10"]+'''</td>
            <td>'''+di["edsuNonMSOpenLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["edsuNonMSOpenCriticalGreaterThan10"])+int(di["edsuNonMSOpenHighGreaterThan10"])+int(di["edsuNonMSOpenAverageGreaterThan10"])+int(di["edsuNonMSOpenLowGreaterThan10"]))+'''</td>
        </tr>

        <tr>
            <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["edsuMSOpenCriticalGreaterThan50"]+'''</td>
            <td >'''+di["edsuMSOpenHighGreaterThan50"]+'''</td>
            <td >'''+di["edsuMSOpenAverageGreaterThan50"]+'''</td>
            <td >'''+di["edsuMSOpenLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["edsuMSOpenCriticalGreaterThan50"])+int(di["edsuMSOpenHighGreaterThan50"])+int(di["edsuMSOpenAverageGreaterThan50"])+int(di["edsuMSOpenLowGreaterThan50"]))+'''</td>
             <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["edsuNonMSOpenCriticalGreaterThan50"]+'''</td>
            <td >'''+di["edsuNonMSOpenHighGreaterThan50"]+'''</td>
            <td >'''+di["edsuNonMSOpenAverageGreaterThan50"]+'''</td>
            <td >'''+di["edsuNonMSOpenLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["edsuNonMSOpenCriticalGreaterThan50"])+int(di["edsuNonMSOpenHighGreaterThan50"])+int(di["edsuNonMSOpenAverageGreaterThan50"])+int(di["edsuNonMSOpenLowGreaterThan50"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["edsuMSOpenCriticalGreaterThan100"]+'''</td>
            <td >'''+di["edsuMSOpenHighGreaterThan100"]+'''</td>
            <td >'''+di["edsuMSOpenAverageGreaterThan100"]+'''</td>
            <td >'''+di["edsuMSOpenLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["edsuMSOpenCriticalGreaterThan100"])+int(di["edsuMSOpenHighGreaterThan100"])+int(di["edsuMSOpenAverageGreaterThan100"])+int(di["edsuMSOpenLowGreaterThan100"]))+'''</td>
             <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["edsuNonMSOpenCriticalGreaterThan100"]+'''</td>
            <td >'''+di["edsuNonMSOpenHighGreaterThan100"]+'''</td>
            <td >'''+di["edsuNonMSOpenAverageGreaterThan100"]+'''</td>
            <td >'''+di["edsuNonMSOpenLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["edsuNonMSOpenCriticalGreaterThan100"])+int(di["edsuNonMSOpenHighGreaterThan100"])+int(di["edsuNonMSOpenAverageGreaterThan100"])+int(di["edsuNonMSOpenLowGreaterThan100"]))+'''</td>
        </tr>

        <tr>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Client</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["edsuMSPending_ClientCriticalLessThan10"]+'''</td>
            <td >'''+di["edsuMSPending_ClientHighLessThan10"]+'''</td>
            <td>'''+di["edsuMSPending_ClientAverageLessThan10"]+'''</td>
            <td>'''+di["edsuMSPending_ClientLowLessThan10"]+'''</td>
            <td>'''+str(int(di["edsuMSPending_ClientCriticalLessThan10"])+int(di["edsuMSPending_ClientHighLessThan10"])+int(di["edsuMSPending_ClientAverageLessThan10"])+int(di["edsuMSPending_ClientLowLessThan10"]))+'''</td>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Client</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["edsuNonMSPending_ClientCriticalLessThan10"]+'''</td>
            <td >'''+di["edsuNonMSPending_ClientHighLessThan10"]+'''</td>
            <td>'''+di["edsuNonMSPending_ClientAverageLessThan10"]+'''</td>
            <td>'''+di["edsuNonMSPending_ClientLowLessThan10"]+'''</td>
            <td>'''+str(int(di["edsuNonMSPending_ClientCriticalLessThan10"])+int(di["edsuNonMSPending_ClientHighLessThan10"])+int(di["edsuNonMSPending_ClientAverageLessThan10"])+int(di["edsuNonMSPending_ClientLowLessThan10"]))+'''</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["edsuMSPending_ClientCriticalGreaterThan10"]+'''</td>
            <td >'''+di["edsuMSPending_ClientHighGreaterThan10"]+'''</td>
            <td>'''+di["edsuMSPending_ClientAverageGreaterThan10"]+'''</td>
            <td>'''+di["edsuMSPending_ClientLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["edsuMSPending_ClientCriticalGreaterThan10"])+int(di["edsuMSPending_ClientHighGreaterThan10"])+int(di["edsuMSPending_ClientAverageGreaterThan10"])+int(di["edsuMSPending_ClientLowGreaterThan10"]))+'''</td>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["edsuNonMSPending_ClientCriticalGreaterThan10"]+'''</td>
            <td >'''+di["edsuNonMSPending_ClientHighGreaterThan10"]+'''</td>
            <td>'''+di["edsuNonMSPending_ClientAverageGreaterThan10"]+'''</td>
            <td>'''+di["edsuNonMSPending_ClientLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["edsuNonMSPending_ClientCriticalGreaterThan10"])+int(di["edsuNonMSPending_ClientHighGreaterThan10"])+int(di["edsuNonMSPending_ClientAverageGreaterThan10"])+int(di["edsuNonMSPending_ClientLowGreaterThan10"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["edsuMSPending_ClientCriticalGreaterThan50"]+'''</td>
            <td >'''+di["edsuMSPending_ClientHighGreaterThan50"]+'''</td>
            <td >'''+di["edsuMSPending_ClientAverageGreaterThan50"]+'''</td>
            <td >'''+di["edsuMSPending_ClientLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["edsuMSPending_ClientCriticalGreaterThan50"])+int(di["edsuMSPending_ClientHighGreaterThan50"])+int(di["edsuMSPending_ClientAverageGreaterThan50"])+int(di["edsuMSPending_ClientLowGreaterThan50"]))+'''</td>
             <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["edsuNonMSPending_ClientCriticalGreaterThan50"]+'''</td>
            <td >'''+di["edsuNonMSPending_ClientHighGreaterThan50"]+'''</td>
            <td >'''+di["edsuNonMSPending_ClientAverageGreaterThan50"]+'''</td>
            <td >'''+di["edsuNonMSPending_ClientLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["edsuNonMSPending_ClientCriticalGreaterThan50"])+int(di["edsuNonMSPending_ClientHighGreaterThan50"])+int(di["edsuNonMSPending_ClientAverageGreaterThan50"])+int(di["edsuNonMSPending_ClientLowGreaterThan50"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["edsuMSPending_ClientCriticalGreaterThan100"]+'''</td>
            <td >'''+di["edsuMSPending_ClientHighGreaterThan100"]+'''</td>
            <td >'''+di["edsuMSPending_ClientAverageGreaterThan100"]+'''</td>
            <td >'''+di["edsuMSPending_ClientLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["edsuMSPending_ClientCriticalGreaterThan100"])+int(di["edsuMSPending_ClientHighGreaterThan100"])+int(di["edsuMSPending_ClientAverageGreaterThan100"])+int(di["edsuMSPending_ClientLowGreaterThan100"]))+'''</td>
             <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["edsuNonMSPending_ClientCriticalGreaterThan100"]+'''</td>
            <td >'''+di["edsuNonMSPending_ClientHighGreaterThan100"]+'''</td>
            <td >'''+di["edsuNonMSPending_ClientAverageGreaterThan100"]+'''</td>
            <td >'''+di["edsuNonMSPending_ClientLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["edsuNonMSPending_ClientCriticalGreaterThan100"])+int(di["edsuNonMSPending_ClientHighGreaterThan100"])+int(di["edsuNonMSPending_ClientAverageGreaterThan100"])+int(di["edsuNonMSPending_ClientLowGreaterThan100"]))+'''</td>
        </tr>
        <tr>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Emergency Client</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["edsuMSPending_Emergency_ClientCriticalLessThan10"]+'''</td>
            <td >'''+di["edsuMSPending_Emergency_ClientHighLessThan10"]+'''</td>
            <td>'''+di["edsuMSPending_Emergency_ClientAverageLessThan10"]+'''</td>
            <td>'''+di["edsuMSPending_Emergency_ClientLowLessThan10"]+'''</td>
            <td>'''+str(int(di["edsuMSPending_Emergency_ClientCriticalLessThan10"])+int(di["edsuMSPending_Emergency_ClientHighLessThan10"])+int(di["edsuMSPending_Emergency_ClientAverageLessThan10"])+int(di["edsuMSPending_Emergency_ClientLowLessThan10"]))+'''</td>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Emergency Client</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["edsuNonMSPending_Emergency_ClientCriticalLessThan10"]+'''</td>
            <td >'''+di["edsuNonMSPending_Emergency_ClientHighLessThan10"]+'''</td>
            <td>'''+di["edsuNonMSPending_Emergency_ClientAverageLessThan10"]+'''</td>
            <td>'''+di["edsuNonMSPending_Emergency_ClientLowLessThan10"]+'''</td>
            <td>'''+str(int(di["edsuNonMSPending_Emergency_ClientCriticalLessThan10"])+int(di["edsuNonMSPending_Emergency_ClientHighLessThan10"])+int(di["edsuNonMSPending_Emergency_ClientAverageLessThan10"])+int(di["edsuNonMSPending_Emergency_ClientLowLessThan10"]))+'''</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["edsuMSPending_Emergency_ClientCriticalGreaterThan10"]+'''</td>
            <td >'''+di["edsuMSPending_Emergency_ClientHighGreaterThan10"]+'''</td>
            <td>'''+di["edsuMSPending_Emergency_ClientAverageGreaterThan10"]+'''</td>
            <td>'''+di["edsuMSPending_Emergency_ClientLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["edsuMSPending_Emergency_ClientCriticalGreaterThan10"])+int(di["edsuMSPending_Emergency_ClientHighGreaterThan10"])+int(di["edsuMSPending_Emergency_ClientAverageGreaterThan10"])+int(di["edsuMSPending_Emergency_ClientLowGreaterThan10"]))+'''</td>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["edsuNonMSPending_Emergency_ClientCriticalGreaterThan10"]+'''</td>
            <td >'''+di["edsuNonMSPending_Emergency_ClientHighGreaterThan10"]+'''</td>
            <td>'''+di["edsuNonMSPending_Emergency_ClientAverageGreaterThan10"]+'''</td>
            <td>'''+di["edsuNonMSPending_Emergency_ClientLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["edsuNonMSPending_Emergency_ClientCriticalGreaterThan10"])+int(di["edsuNonMSPending_Emergency_ClientHighGreaterThan10"])+int(di["edsuNonMSPending_Emergency_ClientAverageGreaterThan10"])+int(di["edsuNonMSPending_Emergency_ClientLowGreaterThan10"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["edsuMSPending_Emergency_ClientCriticalGreaterThan50"]+'''</td>
            <td >'''+di["edsuMSPending_Emergency_ClientHighGreaterThan50"]+'''</td>
            <td >'''+di["edsuMSPending_Emergency_ClientAverageGreaterThan50"]+'''</td>
            <td >'''+di["edsuMSPending_Emergency_ClientLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["edsuMSPending_Emergency_ClientCriticalGreaterThan50"])+int(di["edsuMSPending_Emergency_ClientHighGreaterThan50"])+int(di["edsuMSPending_Emergency_ClientAverageGreaterThan50"])+int(di["edsuMSPending_Emergency_ClientLowGreaterThan50"]))+'''</td>
             <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["edsuNonMSPending_Emergency_ClientCriticalGreaterThan50"]+'''</td>
            <td >'''+di["edsuNonMSPending_Emergency_ClientHighGreaterThan50"]+'''</td>
            <td >'''+di["edsuNonMSPending_Emergency_ClientAverageGreaterThan50"]+'''</td>
            <td >'''+di["edsuNonMSPending_Emergency_ClientLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["edsuNonMSPending_Emergency_ClientCriticalGreaterThan50"])+int(di["edsuNonMSPending_Emergency_ClientHighGreaterThan50"])+int(di["edsuNonMSPending_Emergency_ClientAverageGreaterThan50"])+int(di["edsuNonMSPending_Emergency_ClientLowGreaterThan50"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["edsuMSPending_Emergency_ClientCriticalGreaterThan100"]+'''</td>
            <td >'''+di["edsuMSPending_Emergency_ClientHighGreaterThan100"]+'''</td>
            <td >'''+di["edsuMSPending_Emergency_ClientAverageGreaterThan100"]+'''</td>
            <td >'''+di["edsuMSPending_Emergency_ClientLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["edsuMSPending_Emergency_ClientCriticalGreaterThan100"])+int(di["edsuMSPending_Emergency_ClientHighGreaterThan100"])+int(di["edsuMSPending_Emergency_ClientAverageGreaterThan100"])+int(di["edsuMSPending_Emergency_ClientLowGreaterThan100"]))+'''</td>
             <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["edsuNonMSPending_Emergency_ClientCriticalGreaterThan100"]+'''</td>
            <td >'''+di["edsuNonMSPending_Emergency_ClientHighGreaterThan100"]+'''</td>
            <td >'''+di["edsuNonMSPending_Emergency_ClientAverageGreaterThan100"]+'''</td>
            <td >'''+di["edsuNonMSPending_Emergency_ClientLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["edsuNonMSPending_Emergency_ClientCriticalGreaterThan100"])+int(di["edsuNonMSPending_Emergency_ClientHighGreaterThan100"])+int(di["edsuNonMSPending_Emergency_ClientAverageGreaterThan100"])+int(di["edsuNonMSPending_Emergency_ClientLowGreaterThan100"]))+'''</td>
        </tr>
        <tr>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Vendor</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["edsuMSPending_VendorCriticalLessThan10"]+'''</td>
            <td >'''+di["edsuMSPending_VendorHighLessThan10"]+'''</td>
            <td>'''+di["edsuMSPending_VendorAverageLessThan10"]+'''</td>
            <td>'''+di["edsuMSPending_VendorLowLessThan10"]+'''</td>
            <td>'''+str(int(di["edsuMSPending_VendorCriticalLessThan10"])+int(di["edsuMSPending_VendorHighLessThan10"])+int(di["edsuMSPending_VendorAverageLessThan10"])+int(di["edsuMSPending_VendorLowLessThan10"]))+'''</td>
            <td rowspan="4" bgcolor="#ADD8E6">Pending Vendor</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["edsuNonMSPending_VendorCriticalLessThan10"]+'''</td>
            <td >'''+di["edsuNonMSPending_VendorHighLessThan10"]+'''</td>
            <td>'''+di["edsuNonMSPending_VendorAverageLessThan10"]+'''</td>
            <td>'''+di["edsuNonMSPending_VendorLowLessThan10"]+'''</td>
            <td>'''+str(int(di["edsuNonMSPending_VendorCriticalLessThan10"])+int(di["edsuNonMSPending_VendorHighLessThan10"])+int(di["edsuNonMSPending_VendorAverageLessThan10"])+int(di["edsuNonMSPending_VendorLowLessThan10"]))+'''</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["edsuMSPending_VendorCriticalGreaterThan10"]+'''</td>
            <td >'''+di["edsuMSPending_VendorHighGreaterThan10"]+'''</td>
            <td>'''+di["edsuMSPending_VendorAverageGreaterThan10"]+'''</td>
            <td>'''+di["edsuMSPending_VendorLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["edsuMSPending_VendorCriticalGreaterThan10"])+int(di["edsuMSPending_VendorHighGreaterThan10"])+int(di["edsuMSPending_VendorAverageGreaterThan10"])+int(di["edsuMSPending_VendorLowGreaterThan10"]))+'''</td>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["edsuNonMSPending_VendorCriticalGreaterThan10"]+'''</td>
            <td >'''+di["edsuNonMSPending_VendorHighGreaterThan10"]+'''</td>
            <td>'''+di["edsuNonMSPending_VendorAverageGreaterThan10"]+'''</td>
            <td>'''+di["edsuNonMSPending_VendorLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["edsuNonMSPending_VendorCriticalGreaterThan10"])+int(di["edsuNonMSPending_VendorHighGreaterThan10"])+int(di["edsuNonMSPending_VendorAverageGreaterThan10"])+int(di["edsuNonMSPending_VendorLowGreaterThan10"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["edsuMSPending_VendorCriticalGreaterThan50"]+'''</td>
            <td >'''+di["edsuMSPending_VendorHighGreaterThan50"]+'''</td>
            <td >'''+di["edsuMSPending_VendorAverageGreaterThan50"]+'''</td>
            <td >'''+di["edsuMSPending_VendorLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["edsuMSPending_VendorCriticalGreaterThan50"])+int(di["edsuMSPending_VendorHighGreaterThan50"])+int(di["edsuMSPending_VendorAverageGreaterThan50"])+int(di["edsuMSPending_VendorLowGreaterThan50"]))+'''</td>
             <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["edsuNonMSPending_VendorCriticalGreaterThan50"]+'''</td>
            <td >'''+di["edsuNonMSPending_VendorHighGreaterThan50"]+'''</td>
            <td >'''+di["edsuNonMSPending_VendorAverageGreaterThan50"]+'''</td>
            <td >'''+di["edsuNonMSPending_VendorLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["edsuNonMSPending_VendorCriticalGreaterThan50"])+int(di["edsuNonMSPending_VendorHighGreaterThan50"])+int(di["edsuNonMSPending_VendorAverageGreaterThan50"])+int(di["edsuNonMSPending_VendorLowGreaterThan50"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["edsuMSPending_VendorCriticalGreaterThan100"]+'''</td>
            <td >'''+di["edsuMSPending_VendorHighGreaterThan100"]+'''</td>
            <td >'''+di["edsuMSPending_VendorAverageGreaterThan100"]+'''</td>
            <td >'''+di["edsuMSPending_VendorLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["edsuMSPending_VendorCriticalGreaterThan100"])+int(di["edsuMSPending_VendorHighGreaterThan100"])+int(di["edsuMSPending_VendorAverageGreaterThan100"])+int(di["edsuMSPending_VendorLowGreaterThan100"]))+'''</td>
             <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["edsuNonMSPending_VendorCriticalGreaterThan100"]+'''</td>
            <td >'''+di["edsuNonMSPending_VendorHighGreaterThan100"]+'''</td>
            <td >'''+di["edsuNonMSPending_VendorAverageGreaterThan100"]+'''</td>
            <td >'''+di["edsuNonMSPending_VendorLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["edsuNonMSPending_VendorCriticalGreaterThan100"])+int(di["edsuNonMSPending_VendorHighGreaterThan100"])+int(di["edsuNonMSPending_VendorAverageGreaterThan100"])+int(di["edsuNonMSPending_VendorLowGreaterThan100"]))+'''</td>
        </tr>
        <tr>
            <td rowspan="4" bgcolor="#ADD8E6">Work In Progress</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["edsuMSWork_In_ProgressCriticalLessThan10"]+'''</td>
            <td >'''+di["edsuMSWork_In_ProgressHighLessThan10"]+'''</td>
            <td>'''+di["edsuMSWork_In_ProgressAverageLessThan10"]+'''</td>
            <td>'''+di["edsuMSWork_In_ProgressLowLessThan10"]+'''</td>
            <td>'''+str(int(di["edsuMSWork_In_ProgressCriticalLessThan10"])+int(di["edsuMSWork_In_ProgressHighLessThan10"])+int(di["edsuMSWork_In_ProgressAverageLessThan10"])+int(di["edsuMSWork_In_ProgressLowLessThan10"]))+'''</td>
            <td rowspan="4" bgcolor="#ADD8E6">Work In Progress</td>
            <td bgcolor="#ffcccb"><=10</td>
            <td >'''+di["edsuNonMSWork_In_ProgressCriticalLessThan10"]+'''</td>
            <td >'''+di["edsuNonMSWork_In_ProgressHighLessThan10"]+'''</td>
            <td>'''+di["edsuNonMSWork_In_ProgressAverageLessThan10"]+'''</td>
            <td>'''+di["edsuNonMSWork_In_ProgressLowLessThan10"]+'''</td>
            <td>'''+str(int(di["edsuNonMSWork_In_ProgressCriticalLessThan10"])+int(di["edsuNonMSWork_In_ProgressHighLessThan10"])+int(di["edsuNonMSWork_In_ProgressAverageLessThan10"])+int(di["edsuNonMSWork_In_ProgressLowLessThan10"]))+'''</td>
        </tr>
        <tr>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["edsuMSWork_In_ProgressCriticalGreaterThan10"]+'''</td>
            <td >'''+di["edsuMSWork_In_ProgressHighGreaterThan10"]+'''</td>
            <td>'''+di["edsuMSWork_In_ProgressAverageGreaterThan10"]+'''</td>
            <td>'''+di["edsuMSWork_In_ProgressLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["edsuMSWork_In_ProgressCriticalGreaterThan10"])+int(di["edsuMSWork_In_ProgressHighGreaterThan10"])+int(di["edsuMSWork_In_ProgressAverageGreaterThan10"])+int(di["edsuMSWork_In_ProgressLowGreaterThan10"]))+'''</td>
            <td bgcolor="#ffcccb">>10</td>
            <td >'''+di["edsuNonMSWork_In_ProgressCriticalGreaterThan10"]+'''</td>
            <td >'''+di["edsuNonMSWork_In_ProgressHighGreaterThan10"]+'''</td>
            <td>'''+di["edsuNonMSWork_In_ProgressAverageGreaterThan10"]+'''</td>
            <td>'''+di["edsuNonMSWork_In_ProgressLowGreaterThan10"]+'''</td>
            <td>'''+str(int(di["edsuNonMSWork_In_ProgressCriticalGreaterThan10"])+int(di["edsuNonMSWork_In_ProgressHighGreaterThan10"])+int(di["edsuNonMSWork_In_ProgressAverageGreaterThan10"])+int(di["edsuNonMSWork_In_ProgressLowGreaterThan10"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["edsuMSWork_In_ProgressCriticalGreaterThan50"]+'''</td>
            <td >'''+di["edsuMSWork_In_ProgressHighGreaterThan50"]+'''</td>
            <td >'''+di["edsuMSWork_In_ProgressAverageGreaterThan50"]+'''</td>
            <td >'''+di["edsuMSWork_In_ProgressLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["edsuMSWork_In_ProgressCriticalGreaterThan50"])+int(di["edsuMSWork_In_ProgressHighGreaterThan50"])+int(di["edsuMSWork_In_ProgressAverageGreaterThan50"])+int(di["edsuMSWork_In_ProgressLowGreaterThan50"]))+'''</td>
             <td  bgcolor="#ffcccb">>50</td>
            <td >'''+di["edsuNonMSWork_In_ProgressCriticalGreaterThan50"]+'''</td>
            <td >'''+di["edsuNonMSWork_In_ProgressHighGreaterThan50"]+'''</td>
            <td >'''+di["edsuNonMSWork_In_ProgressAverageGreaterThan50"]+'''</td>
            <td >'''+di["edsuNonMSWork_In_ProgressLowGreaterThan50"]+'''</td>
            <td >'''+str(int(di["edsuNonMSWork_In_ProgressCriticalGreaterThan50"])+int(di["edsuNonMSWork_In_ProgressHighGreaterThan50"])+int(di["edsuNonMSWork_In_ProgressAverageGreaterThan50"])+int(di["edsuNonMSWork_In_ProgressLowGreaterThan50"]))+'''</td>
        </tr>
        <tr>
            <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["edsuMSWork_In_ProgressCriticalGreaterThan100"]+'''</td>
            <td >'''+di["edsuMSWork_In_ProgressHighGreaterThan100"]+'''</td>
            <td >'''+di["edsuMSWork_In_ProgressAverageGreaterThan100"]+'''</td>
            <td >'''+di["edsuMSWork_In_ProgressLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["edsuMSWork_In_ProgressCriticalGreaterThan100"])+int(di["edsuMSWork_In_ProgressHighGreaterThan100"])+int(di["edsuMSWork_In_ProgressAverageGreaterThan100"])+int(di["edsuMSWork_In_ProgressLowGreaterThan100"]))+'''</td>
             <td  bgcolor="#ffcccb">>100</td>
            <td >'''+di["edsuNonMSWork_In_ProgressCriticalGreaterThan100"]+'''</td>
            <td >'''+di["edsuNonMSWork_In_ProgressHighGreaterThan100"]+'''</td>
            <td >'''+di["edsuNonMSWork_In_ProgressAverageGreaterThan100"]+'''</td>
            <td >'''+di["edsuNonMSWork_In_ProgressLowGreaterThan100"]+'''</td>
            <td >'''+str(int(di["edsuNonMSWork_In_ProgressCriticalGreaterThan100"])+int(di["edsuNonMSWork_In_ProgressHighGreaterThan100"])+int(di["edsuNonMSWork_In_ProgressAverageGreaterThan100"])+int(di["edsuNonMSWork_In_ProgressLowGreaterThan100"]))+'''</td>
        </tr>
        
    </table>
    </div>
  
    
<br /><font face='Cambria'>Regards </a></font><br/>
<br /><font face='Cambria'>3M Automation Center Team </a></font><br/>
<br>
<br>
<br/><img src='cid:image3'<br/>
</body>
</html>'''
From ='USSACPrd@mmm.com' 
To = sys.argv[4]
cc=sys.argv[5]
#bcc=""

msgRoot = MIMEMultipart('related')
msgRoot['Subject'] = sys.argv[6]+"-"+sys.argv[8]
msgRoot['From'] = From
msgRoot['Cc']=cc+","+sys.argv[9]
msgRoot['To'] = To
#msgRoot['Bcc']=bcc
msgRoot.preamble = '====================================================='
msgAlternative = MIMEMultipart('alternative')
msgRoot.attach(msgAlternative)
msgText = MIMEText('Please find ')
msgAlternative.attach(msgText)
msgText = MIMEText(html_file, 'html')
msgAlternative.attach(msgText)
msgAlternative.attach(msgText)
fp = open(r"\\acprd01\3M_CAC\EDI_Ageing\head.png",'rb')
#fp2 = open(sys.argv[7], 'rb')#"//acdev01/3M_CAC/IPM_FSM/Mail_elements/new.png"
fp3 = open(r"\\acprd01\3M_CAC\EDI_Ageing\footer.png",'rb')
msgImage = MIMEImage(fp.read())
#msgImage1 = MIMEImage(fp2.read())
msgImage2 = MIMEImage(fp3.read())
fp.close()
fp3.close()
msgImage.add_header('Content-ID', '<image1>')
msgImage2.add_header('Content-ID', '<image3>')
msgRoot.attach(msgImage)
msgRoot.attach(msgImage2)
filepaths=[sys.argv[7]]
for f in filepaths :
    with open(f,"rb") as file:
        part=MIMEApplication(file.read(),Name=basename(f))
        part["Content-Disposition"]='attachment;filename="%s"'%basename(f)
        msgRoot.attach(part)
smtp = smtplib.SMTP()
smtp.connect("mailserv.mmm.com")
#smtp.sendmail(From,To, msgRoot.as_string())
smtp.send_message(msgRoot)
smtp.quit()
print("Email is sent successfully")
