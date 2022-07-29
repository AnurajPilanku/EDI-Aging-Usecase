'''
Created by     : AnurajPilanku
Code utility   : remove rows based on a value in a column
Use case        : EDI Use case
'''
import datetime
import openpyxl
import sys

today=datetime.date.today()
year=today.year
month=str(datetime.datetime.now().strftime("%B"))
time=str(datetime.datetime.now())[10:][:6]
day=today.day

wb=openpyxl.load_workbook(sys.argv[1])
ws=wb.active
ediCC=list()
for i in range(2,ws.max_row+1):
    if ws['A'+str(i)].value not in [None,"NULL",""," "]:
        ediCC.append(ws['A'+str(i)].value.strip())
    if ws['B' + str(i)].value not in [None, "NULL", "", " "]:
        ediCC.append(ws['B' + str(i)].value.strip())
    if ws['C' + str(i)].value not in [None, "NULL", "", " "]:
        ediCC.append(ws['C' + str(i)].value.strip())
uniqEDIcc=",".join(list(set(ediCC)))

paths={}
paths['ediexcelcarboncopy']=uniqEDIcc
paths["todaysDate"]=str(day)+" "+str(datetime.datetime.now().strftime("%B"))+" "+str(year)+" "+time

output = {'output':paths,'additional_attributes':paths}
sys.stdout.write(str(output)+'\n')

